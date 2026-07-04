#!/usr/bin/env python3
"""
SPPG Report Generator — Satuan Pelayanan Pemenuhan Gizi
Generate Excel laporan bulanan SPPG
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
import calendar
import os

# ── Color palette ──
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

FONT_TITLE = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
FONT_HEADER = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
FONT_SUBHEADER = Font(name="Calibri", bold=True, size=10, color="1F4E79")
FONT_BODY = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", bold=True, size=10)
FONT_MONEY = Font(name="Calibri", size=10)
FONT_TOTAL = Font(name="Calibri", bold=True, size=11, color="1F4E79")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
MONEY_FORMAT = '#,##0'

# ── Sample data ──
MONTH = 7   # Juli
YEAR = 2026
LOCATION = "Kecamatan Sukamaju, Kota Bandung"

PENERIMA = [
    {"id": "SPPG-001", "nama": "Ani Rahmawati", "usia": 35, "alamat": "RW 01 Kel. Sukamaju", "kategori": "Ibu Hamil", "status": "Aktif"},
    {"id": "SPPG-002", "nama": "Budi Santoso",   "usia": 42, "alamat": "RW 01 Kel. Sukamaju", "kategori": "Balita",   "status": "Aktif"},
    {"id": "SPPG-003", "nama": "Citra Dewi",     "usia": 28, "alamat": "RW 02 Kel. Sukamaju", "kategori": "Ibu Menyusui", "status": "Aktif"},
    {"id": "SPPG-004", "nama": "Dodi Firmansyah", "usia": 3,  "alamat": "RW 02 Kel. Sukamaju", "kategori": "Balita",   "status": "Aktif"},
    {"id": "SPPG-005", "nama": "Eka Putri",      "usia": 26, "alamat": "RW 03 Kel. Sukamaju", "kategori": "Ibu Hamil", "status": "Aktif"},
    {"id": "SPPG-006", "nama": "Farhan Hakim",   "usia": 4,  "alamat": "RW 03 Kel. Sukamaju", "kategori": "Balita",   "status": "Aktif"},
    {"id": "SPPG-007", "nama": "Gita Permata",   "usia": 31, "alamat": "RW 01 Kel. Sukamaju", "kategori": "Ibu Menyusui", "status": "Aktif"},
    {"id": "SPPG-008", "nama": "Hendra Gunawan", "usia": 2,  "alamat": "RW 04 Kel. Sukamaju", "kategori": "Balita",   "status": "Aktif"},
    {"id": "SPPG-009", "nama": "Intan Nurhaliza", "usia": 27, "alamat": "RW 04 Kel. Sukamaju", "kategori": "Ibu Hamil", "status": "Aktif"},
    {"id": "SPPG-010", "nama": "Joko Prasetyo",  "usia": 5,  "alamat": "RW 05 Kel. Sukamaju", "kategori": "Balita",   "status": "Aktif"},
]

MENU = [
    {"kode": "M-001", "nama": "Nasi + Ayam Goreng + Sayur + Susu",         "jenis": "Makan Siang", "biaya": 15000},
    {"kode": "M-002", "nama": "Nasi + Ikan Bakar + Sayur + Buah",           "jenis": "Makan Siang", "biaya": 18000},
    {"kode": "M-003", "nama": "Bubur Kacang Hijau + Telur + Susu",         "jenis": "Sarapan",     "biaya": 12000},
    {"kode": "M-004", "nama": "Nasi + Tumis Tahu Tempe + Ikan + Sayur",    "jenis": "Makan Siang", "biaya": 16000},
    {"kode": "M-005", "nama": "Susu + Biskuit + Buah",                      "jenis": "Snack",       "biaya": 8000},
]


def style_header_row(ws, row, max_col, fill=HEADER_FILL, font=FONT_HEADER):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT_BODY
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if (r - start_row) % 2 == 1:
                cell.fill = LIGHT_GRAY


def auto_width(ws, max_col, min_width=10, max_width=40):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        lengths = []
        for cell in ws[letter]:
            if cell.value:
                lengths.append(len(str(cell.value)))
        best = max(lengths) + 3 if lengths else min_width
        ws.column_dimensions[letter].width = min(max(best, min_width), max_width)


# ═══════════════════════════════════════════
#  MAIN GENERATOR
# ═══════════════════════════════════════════
def generate():
    wb = openpyxl.Workbook()

    bulan_nama = calendar.month_name[MONTH]
    filename = f"SPPG_Report_{bulan_nama}_{YEAR}.xlsx"
    filepath = os.path.join(os.path.dirname(__file__), filename)

    # ── Sheet 1: Ringkasan ──
    ws1 = wb.active
    ws1.title = "Ringkasan"
    ws1.sheet_properties.tabColor = "1F4E79"

    # Title
    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value = f"LAPORAN BULANAN SPPG — {bulan_nama.upper()} {YEAR}"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    title_cell.alignment = CENTER

    ws1.merge_cells("A2:F2")
    ws1["A2"].value = f"Lokasi: {LOCATION}"
    ws1["A2"].font = Font(name="Calibri", size=11, color="555555")
    ws1["A2"].alignment = CENTER

    # Key metrics
    total_penerima = len(PENERIMA)
    aktif = sum(1 for p in PENERIMA if p["status"] == "Aktif")
    hari_kerja = 22  # asumsi Sen-Jum
    rata_menu = sum(m["biaya"] for m in MENU) / len(MENU)
    total_biaya_per_hari = aktif * rata_menu
    total_bulanan = total_biaya_per_hari * hari_kerja

    metrics = [
        ("Total Penerima Manfaat", aktif, "Orang", "👥"),
        ("Total Hari Distribusi", hari_kerja, "Hari", "📅"),
        ("Rata-rata Biaya/Menu/Hari", f"Rp {rata_menu:,.0f}", "", "💰"),
        ("Estimasi Biaya Bulanan", f"Rp {total_bulanan:,.0f}", "", "💵"),
        ("Kategori Aktif", "3", "Ibu Hamil / Menyusui / Balita", "👶"),
    ]

    ws1["A4"].value = "📊 INDIKATOR UTAMA"
    ws1["A4"].font = FONT_SUBHEADER

    for i, (label, val, unit, emoji) in enumerate(metrics):
        row = 5 + i
        ws1[f"A{row}"].value = f"{emoji} {label}"
        ws1[f"A{row}"].font = FONT_BOLD
        ws1[f"B{row}"].value = val
        ws1[f"B{row}"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        ws1[f"C{row}"].value = unit
        ws1[f"C{row}"].font = FONT_BODY
        for c in range(1, 4):
            ws1.cell(row=row, column=c).border = THIN_BORDER
            ws1.cell(row=row, column=c).alignment = LEFT

    # Breakdown by kategori
    ws1[f"A{11}"].value = "📋 RINCIAN PER KATEGORI"
    ws1[f"A{11}"].font = FONT_SUBHEADER

    kat_header = ["Kategori", "Jumlah", "Persentase"]
    for j, h in enumerate(kat_header, 1):
        ws1.cell(row=12, column=j, value=h)
    style_header_row(ws1, 12, 3)

    kategori_count = {}
    for p in PENERIMA:
        kategori_count[p["kategori"]] = kategori_count.get(p["kategori"], 0) + 1

    for i, (kat, jml) in enumerate(sorted(kategori_count.items())):
        row = 13 + i
        ws1.cell(row=row, column=1, value=kat)
        ws1.cell(row=row, column=2, value=jml)
        ws1.cell(row=row, column=3, value=f"{jml/aktif*100:.0f}%")
    style_data_rows(ws1, 13, 12 + len(kategori_count), 3)

    # Summary by age group
    ws1[f"A{18}"].value = "📈 DATA USIA PENERIMA"
    ws1[f"A{18}"].font = FONT_SUBHEADER

    usia_header = ["Kategori Usia", "Jumlah"]
    for j, h in enumerate(usia_header, 1):
        ws1.cell(row=19, column=j, value=h)
    style_header_row(ws1, 19, 2)

    dewasa = sum(1 for p in PENERIMA if p["usia"] >= 18)
    balita = sum(1 for p in PENERIMA if p["usia"] < 6)
    anak = sum(1 for p in PENERIMA if 6 <= p["usia"] < 18)

    usia_data = [("Balita (< 6 tahun)", balita), ("Anak-anak (6-17)", anak), ("Dewasa (≥ 18)", dewasa)]
    for i, (kat, jml) in enumerate(usia_data):
        ws1.cell(row=20 + i, column=1, value=kat)
        ws1.cell(row=20 + i, column=2, value=jml)
    style_data_rows(ws1, 20, 22, 2)

    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 30

    # ── Sheet 2: Data Penerima ──
    ws2 = wb.create_sheet("Data Penerima")
    ws2.sheet_properties.tabColor = "2E75B6"

    ws2.merge_cells("A1:E1")
    ws2["A1"].value = "DATA PENERIMA MANFAAT SPPG"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws2["A1"].alignment = CENTER

    headers2 = ["ID Penerima", "Nama Lengkap", "Usia", "Alamat", "Kategori", "Status"]
    for j, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=j, value=h)
    style_header_row(ws2, 3, len(headers2))

    for i, p in enumerate(PENERIMA):
        row = 4 + i
        ws2.cell(row=row, column=1, value=p["id"])
        ws2.cell(row=row, column=2, value=p["nama"])
        ws2.cell(row=row, column=3, value=p["usia"])
        ws2.cell(row=row, column=4, value=p["alamat"])
        ws2.cell(row=row, column=5, value=p["kategori"])
        ws2.cell(row=row, column=6, value=p["status"])
    style_data_rows(ws2, 4, 3 + len(PENERIMA), len(headers2))
    auto_width(ws2, len(headers2))

    # ── Sheet 3: Menu Makanan ──
    ws3 = wb.create_sheet("Menu Makanan")
    ws3.sheet_properties.tabColor = "548235"

    ws3.merge_cells("A1:D1")
    ws3["A1"].value = "DATABASE MENU MAKANAN SPPG"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws3["A1"].alignment = CENTER

    headers3 = ["Kode Menu", "Nama Menu", "Jenis", "Biaya Satuan (Rp)"]
    for j, h in enumerate(headers3, 1):
        ws3.cell(row=3, column=j, value=h)
    style_header_row(ws3, 3, len(headers3))

    for i, m in enumerate(MENU):
        row = 4 + i
        ws3.cell(row=row, column=1, value=m["kode"])
        ws3.cell(row=row, column=2, value=m["nama"])
        ws3.cell(row=row, column=3, value=m["jenis"])
        ws3.cell(row=row, column=4, value=m["biaya"])
        ws3.cell(row=row, column=4).number_format = MONEY_FORMAT
    style_data_rows(ws3, 4, 3 + len(MENU), len(headers3))
    auto_width(ws3, len(headers3))

    # ── Sheet 4: Distribusi Harian ──
    ws4 = wb.create_sheet("Distribusi Harian")
    ws4.sheet_properties.tabColor = "BF8F00"

    ws4.merge_cells("A1:H1")
    ws4["A1"].value = f"LOG DISTRIBUSI HARIAN — {bulan_nama} {YEAR}"
    ws4["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws4["A1"].alignment = CENTER

    headers4 = ["Tanggal", "Hari", "ID Penerima", "Nama Penerima", "Menu", "Jumlah", "Biaya Satuan", "Total Biaya"]
    for j, h in enumerate(headers4, 1):
        ws4.cell(row=3, column=j, value=h)
    style_header_row(ws4, 3, len(headers4))

    # Generate daily log for first week as sample
    hari_nama = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
    first_day = date(YEAR, MONTH, 1)
    row = 4
    total_all = 0
    sample_days = min(hari_kerja, 7)
    for d in range(sample_days):
        tgl = first_day + timedelta(days=d)
        if tgl.weekday() >= 5:  # skip weekend
            continue
        h_nama = hari_nama[tgl.weekday()]
        # Assign menu rotating
        menu_today = MENU[d % len(MENU)]
        for p in PENERIMA[:4]:  # sample first 4 recipients
            ws4.cell(row=row, column=1, value=tgl.strftime("%d/%m/%Y"))
            ws4.cell(row=row, column=2, value=h_nama)
            ws4.cell(row=row, column=3, value=p["id"])
            ws4.cell(row=row, column=4, value=p["nama"])
            ws4.cell(row=row, column=5, value=menu_today["kode"] + " - " + menu_today["nama"])
            ws4.cell(row=row, column=6, value=1)
            ws4.cell(row=row, column=7, value=menu_today["biaya"])
            ws4.cell(row=row, column=8, value=menu_today["biaya"])
            ws4.cell(row=row, column=8).number_format = MONEY_FORMAT
            total_all += menu_today["biaya"]
            row += 1

    # Total row
    ws4.cell(row=row, column=1, value="TOTAL")
    ws4.cell(row=row, column=1).font = FONT_TOTAL
    ws4.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    ws4.cell(row=row, column=8, value=total_all)
    ws4.cell(row=row, column=8).font = FONT_TOTAL
    ws4.cell(row=row, column=8).number_format = MONEY_FORMAT
    ws4.cell(row=row, column=8).border = THIN_BORDER

    style_data_rows(ws4, 4, row - 1, len(headers4))
    auto_width(ws4, len(headers4))
    ws4.column_dimensions["E"].width = 45

    # ── Sheet 5: Rekap Bulanan ──
    ws5 = wb.create_sheet("Rekap Bulanan")
    ws5.sheet_properties.tabColor = "843C0C"

    ws5.merge_cells("A1:F1")
    ws5["A1"].value = f"REKAP BULANAN PER PENERIMA — {bulan_nama} {YEAR}"
    ws5["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws5["A1"].alignment = CENTER

    headers5 = ["ID", "Nama", "Kategori", "Total Hari", "Rata-rata Biaya/Hari", "Total Bulanan"]
    for j, h in enumerate(headers5, 1):
        ws5.cell(row=3, column=j, value=h)
    style_header_row(ws5, 3, len(headers5))

    grand_total = 0
    for i, p in enumerate(PENERIMA[:6]):  # sample
        row = 4 + i
        biaya_per_hari = rata_menu
        total = biaya_per_hari * hari_kerja
        grand_total += total
        ws5.cell(row=row, column=1, value=p["id"])
        ws5.cell(row=row, column=2, value=p["nama"])
        ws5.cell(row=row, column=3, value=p["kategori"])
        ws5.cell(row=row, column=4, value=hari_kerja)
        ws5.cell(row=row, column=5, value=biaya_per_hari)
        ws5.cell(row=row, column=5).number_format = MONEY_FORMAT
        ws5.cell(row=row, column=6, value=total)
        ws5.cell(row=row, column=6).number_format = MONEY_FORMAT
    style_data_rows(ws5, 4, 3 + min(len(PENERIMA), 6), len(headers5))

    total_row = 4 + min(len(PENERIMA), 6)
    ws5.cell(row=total_row, column=1, value="GRAND TOTAL")
    ws5.cell(row=total_row, column=1).font = FONT_TOTAL
    ws5.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=5)
    ws5.cell(row=total_row, column=6, value=grand_total)
    ws5.cell(row=total_row, column=6).font = FONT_TOTAL
    ws5.cell(row=total_row, column=6).number_format = MONEY_FORMAT
    for c in range(1, len(headers5) + 1):
        ws5.cell(row=total_row, column=c).border = THIN_BORDER

    auto_width(ws5, len(headers5))

    # ── Sheet 6: Rekap Anggaran ──
    ws6 = wb.create_sheet("Rekap Anggaran")
    ws6.sheet_properties.tabColor = "C00000"

    ws6.merge_cells("A1:E1")
    ws6["A1"].value = f"REKAP ANGGARAN — {bulan_nama} {YEAR}"
    ws6["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws6["A1"].alignment = CENTER

    headers6 = ["Pos Anggaran", "Anggaran (Rp)", "Realisasi (Rp)", "Sisa (Rp)", "% Realisasi"]
    for j, h in enumerate(headers6, 1):
        ws6.cell(row=3, column=j, value=h)
    style_header_row(ws6, 3, len(headers6))

    # Dummy budget data
    anggaran_total = 50_000_000
    realisasi_makanan = total_bulanan * 0.85
    realisasi_operasional = total_bulanan * 0.10
    realisasi_lain = total_bulanan * 0.05
    realisasi_total = realisasi_makanan + realisasi_operasional + realisasi_lain

    budget_items = [
        ("Makanan & Gizi", anggaran_total * 0.80, realisasi_makanan),
        ("Operasional Dapur", anggaran_total * 0.12, realisasi_operasional),
        ("Lain-lain", anggaran_total * 0.08, realisasi_lain),
    ]

    for i, (pos, anggaran, realisasi) in enumerate(budget_items):
        row = 4 + i
        sisa = anggaran - realisasi
        pct = (realisasi / anggaran) * 100 if anggaran > 0 else 0
        ws6.cell(row=row, column=1, value=pos)
        ws6.cell(row=row, column=2, value=anggaran)
        ws6.cell(row=row, column=2).number_format = MONEY_FORMAT
        ws6.cell(row=row, column=3, value=realisasi)
        ws6.cell(row=row, column=3).number_format = MONEY_FORMAT
        ws6.cell(row=row, column=4, value=sisa)
        ws6.cell(row=row, column=4).number_format = MONEY_FORMAT
        ws6.cell(row=row, column=5, value=f"{pct:.0f}%")
        if pct > 95:
            ws6.cell(row=row, column=5).fill = WARN_FILL

    # Total row
    total_r = 4 + len(budget_items)
    ws6.cell(row=total_r, column=1, value="TOTAL ANGGARAN")
    ws6.cell(row=total_r, column=1).font = FONT_TOTAL
    ws6.cell(row=total_r, column=2, value=anggaran_total)
    ws6.cell(row=total_r, column=2).font = FONT_TOTAL
    ws6.cell(row=total_r, column=2).number_format = MONEY_FORMAT
    ws6.cell(row=total_r, column=3, value=realisasi_total)
    ws6.cell(row=total_r, column=3).font = FONT_TOTAL
    ws6.cell(row=total_r, column=3).number_format = MONEY_FORMAT
    ws6.cell(row=total_r, column=4, value=anggaran_total - realisasi_total)
    ws6.cell(row=total_r, column=4).font = FONT_TOTAL
    ws6.cell(row=total_r, column=4).number_format = MONEY_FORMAT
    ws6.cell(row=total_r, column=5, value=f"{realisasi_total/anggaran_total*100:.0f}%")
    ws6.cell(row=total_r, column=5).font = FONT_TOTAL
    for c in range(1, len(headers6) + 1):
        ws6.cell(row=total_r, column=c).border = THIN_BORDER

    style_data_rows(ws6, 4, total_r - 1, len(headers6))
    auto_width(ws6, len(headers6))

    # ── Save ──
    wb.save(filepath)
    print(f"✅ Report generated: {filepath}")
    return filepath


if __name__ == "__main__":
    generate()
