#!/usr/bin/env python3
"""
LAPORAN MINGGUAN SPPG — Weekly Summary Report
Generate from daily JSON files + input analisa & rekomendasi
"""

import sys
import json
from pathlib import Path
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from utils import (
    DATA_DIR, HARIAN_DIR, MINGGUAN_DIR,
    tgl_str, tgl_indo, nama_hari, rupiah, save_json, load_json, init_dirs,
    range_minggu
)

# ── Styling ──
HEADER = PatternFill("solid", fgColor="1F4E79")
SUBHEADER = PatternFill("solid", fgColor="D6E4F0")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
WARN = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="FCE4EC")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
YELLOW_FILL = PatternFill("solid", fgColor="FFF9C4")

FONT_TITLE = Font(name="Calibri", bold=True, size=14, color="1F4E79")
FONT_HEADER = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", bold=True, size=10)
FONT_BODY = Font(name="Calibri", size=10)
FONT_SECTION = Font(name="Calibri", bold=True, size=11, color="1F4E79")

THIN = Border(
    left=Side("thin", "B4C6E7"), right=Side("thin", "B4C6E7"),
    top=Side("thin", "B4C6E7"), bottom=Side("thin", "B4C6E7"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def load_week_data(tgl_start=None, tgl_end=None):
    """Load all daily reports in a date range"""
    if tgl_start is None or tgl_end is None:
        tgl_end = date.today()
        tgl_start = tgl_end - timedelta(days=6)
        # Find Monday
        tgl_start = tgl_end - timedelta(days=tgl_end.weekday())

    if isinstance(tgl_start, str):
        tgl_start = datetime.strptime(tgl_start, "%Y-%m-%d").date()
    if isinstance(tgl_end, str):
        tgl_end = datetime.strptime(tgl_end, "%Y-%m-%d").date()

    days = []
    current = tgl_start
    while current <= tgl_end:
        if current.weekday() < 6:  # Mon-Sat (hari operasional)
            json_path = HARIAN_DIR / f"{current}.json"
            if json_path.exists():
                days.append(load_json(json_path))
            else:
                days.append({
                    "tanggal": str(current),
                    "hari": nama_hari(str(current)),
                    "produksi": 0,
                    "distribusi": 0,
                    "target": 0,
                    "tepat_waktu": False,
                    "kebersihan": "-",
                    "jml_besar": 0, "jml_tendik": 0,
                    "jml_kecil": 0, "jml_3b": 0,
                    "waste_nasi": 0, "waste_sayur": 0,
                    "waste_lauk": 0, "waste_buah": 0,
                    "masalah": "",
                    "tindakan": "",
                    "feedback_pm": "",
                    "catatan": "",
                    "menu_karbohidrat": "", "menu_protein_hewani": "",
                    "menu_protein_nabati": "", "menu_sayur": "", "menu_buah": "",
                })
        current += timedelta(days=1)

    return days, tgl_start, tgl_end


def render_excel(report, filepath):
    """Render full weekly report to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Report"

    ROW = 1

    # ── TITLE ──
    ws.merge_cells("A1:H1")
    ws["A1"].value = f"📊 WEEKLY REPORT — {report['nama_sppg']}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws["A1"].alignment = CENTER
    ROW = 2
    ws.merge_cells(f"A{ROW}:H{ROW}")
    ws[f"A{ROW}"].value = f"Periode: {tgl_indo(report['tgl_start'])} – {tgl_indo(report['tgl_end'])}"
    ws[f"A{ROW}"].font = Font(name="Calibri", size=11, color="555555")
    ws[f"A{ROW}"].alignment = CENTER
    ROW = 4

    # ── SECTION 1: REKAP PRODUKSI & DISTRIBUSI ──
    ws.merge_cells(f"A{ROW}:H{ROW}")
    ws[f"A{ROW}"].value = "🎯 REKAP PRODUKSI & DISTRIBUSI"
    ws[f"A{ROW}"].font = FONT_SECTION
    ws[f"A{ROW}"].fill = SUBHEADER
    ROW += 1

    # Table header
    headers = ["Tanggal", "Hari", "Target", "Produksi", "Distribusi", "Tepat Waktu", "Kebersihan", "Keterangan"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(ROW, c, h)
        cell.font = FONT_HEADER
        cell.fill = HEADER
        cell.alignment = CENTER
        cell.border = THIN
    ROW += 1

    total_target = 0
    total_produksi = 0
    total_distribusi = 0
    tepat_count = 0
    hari_aktif = 0

    for d in report["days"]:
        if d["target"] == 0 and "LIBUR" in d.get("hari", ""):
            ws.cell(ROW, 1, tgl_indo(d["tanggal"]))
            ws.cell(ROW, 2, d.get("hari", ""))
            ws.merge_cells(f"C{ROW}:H{ROW}")
            ws.cell(ROW, 3, "LIBUR NASIONAL / LIBUR")
            for c in range(1, 9):
                ws.cell(ROW, c).font = Font(name="Calibri", italic=True, size=10, color="999999")
                ws.cell(ROW, c).fill = LIGHT_GRAY
                ws.cell(ROW, c).border = THIN
                ws.cell(ROW, c).alignment = CENTER
            ROW += 1
            continue

        hari_aktif += 1
        total_target += d["target"]
        total_produksi += d["produksi"]
        total_distribusi += d["distribusi"]
        if d.get("tepat_waktu", False):
            tepat_count += 1

        ws.cell(ROW, 1, tgl_indo(d["tanggal"])).font = FONT_BODY
        ws.cell(ROW, 2, d.get("hari", "")).font = FONT_BODY
        ws.cell(ROW, 3, d["target"]).font = FONT_BODY
        ws.cell(ROW, 4, d["produksi"]).font = FONT_BODY
        ws.cell(ROW, 5, d["distribusi"]).font = FONT_BODY
        ws.cell(ROW, 6, "✅" if d.get("tepat_waktu", False) else "❌").font = FONT_BODY
        ws.cell(ROW, 7, d.get("kebersihan", "").title() if d.get("kebersihan") else "-").font = FONT_BODY
        ws.cell(ROW, 8, d.get("catatan", "")[:50] if d.get("catatan") else "").font = Font(name="Calibri", size=9, color="666666")
        for c in range(1, 9):
            ws.cell(ROW, c).border = THIN
            ws.cell(ROW, c).alignment = CENTER
        if (ROW % 2) == 0:
            for c in range(1, 9):
                ws.cell(ROW, c).fill = LIGHT_GRAY
        ROW += 1

    # Summary row
    ws.cell(ROW, 1, "TOTAL MINGGUAN").font = FONT_BOLD
    ws.merge_cells(f"A{ROW}:B{ROW}")
    ws.cell(ROW, 3, total_target).font = FONT_BOLD
    ws.cell(ROW, 4, total_produksi).font = FONT_BOLD
    ws.cell(ROW, 5, total_distribusi).font = FONT_BOLD
    pct = (tepat_count / hari_aktif * 100) if hari_aktif else 0
    ws.cell(ROW, 6, f"{pct:.0f}%").font = FONT_BOLD
    for c in range(1, 9):
        ws.cell(ROW, c).border = THIN
        ws.cell(ROW, c).fill = SUBHEADER
        ws.cell(ROW, c).alignment = CENTER
    ROW += 2

    # Key metrics
    metrics = [
        ("Total Produksi", f"{total_produksi:,} Porsi".replace(",", ".")),
        ("Total Distribusi", f"{total_distribusi:,} Porsi".replace(",", ".")),
        ("Ketepatan Waktu", f"{pct:.0f}%"),
        ("Tingkat Pencapaian Produksi", f"{(total_produksi/total_target*100) if total_target else 0:.0f}%"),
        ("Hari Operasional Efektif", f"{hari_aktif} hari"),
    ]
    for label, val in metrics:
        ws.cell(ROW, 1, label).font = FONT_BOLD
        ws.cell(ROW, 1).fill = LIGHT_GRAY
        ws.cell(ROW, 2, val).font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
        for c in range(1, 3):
            ws.cell(ROW, c).border = THIN
            ws.cell(ROW, c).alignment = LEFT
        ROW += 1
    ROW += 1

    # ── SECTION 2: RINCIAN PENERIMA MANFAAT ──
    ws.merge_cells(f"A{ROW}:H{ROW}")
    ws[f"A{ROW}"].value = "👥 RINCIAN PENERIMA MANFAAT MINGGUAN"
    ws[f"A{ROW}"].font = FONT_SECTION
    ws[f"A{ROW}"].fill = SUBHEADER
    ROW += 1

    pm_headers = ["Kategori", "Rata-rata/Hari", "Total Minggu"]
    for c, h in enumerate(pm_headers, 1):
        ws.cell(ROW, c, h).font = FONT_HEADER
        ws.cell(ROW, c).fill = HEADER
        ws.cell(ROW, c).alignment = CENTER
        ws.cell(ROW, c).border = THIN
    ROW += 1

    total_besar = sum(d.get("jml_besar", 0) for d in report["days"])
    total_tendik = sum(d.get("jml_tendik", 0) for d in report["days"])
    total_kecil = sum(d.get("jml_kecil", 0) for d in report["days"])
    total_3b = sum(d.get("jml_3b", 0) for d in report["days"])
    avg_hari = max(hari_aktif, 1)

    pm_data = [
        ("Besar (SD 4-6/SMP/SMA/ATS 9-18)", total_besar // avg_hari, total_besar),
        ("Tendik (Pendidik & Tenaga Kependidikan)", total_tendik // avg_hari, total_tendik),
        ("Kecil (PAUD/TK/SD 1-3/Balita/ATS<9)", total_kecil // avg_hari, total_kecil),
        ("3B (Ibu Hamil/Menyusui/Balita)", total_3b // avg_hari, total_3b),
    ]
    grand_total = total_besar + total_tendik + total_kecil + total_3b
    for label, avg, total in pm_data:
        ws.cell(ROW, 1, label).font = FONT_BODY
        ws.cell(ROW, 2, avg).font = FONT_BODY
        ws.cell(ROW, 3, total).font = FONT_BODY
        for c in range(1, 4):
            ws.cell(ROW, c).border = THIN
            ws.cell(ROW, c).alignment = CENTER
        ROW += 1

    ws.cell(ROW, 1, "TOTAL").font = FONT_BOLD
    ws.cell(ROW, 2, "").font = FONT_BOLD
    ws.cell(ROW, 3, grand_total).font = FONT_BOLD
    for c in range(1, 4):
        ws.cell(ROW, c).border = THIN
        ws.cell(ROW, c).fill = SUBHEADER
        ws.cell(ROW, c).alignment = CENTER
    ROW += 2

    # ── SECTION 3: WASTE FOOD ──
    ws.merge_cells(f"A{ROW}:H{ROW}")
    ws[f"A{ROW}"].value = "🗑️  REKAP WASTE MAKANAN MINGGUAN"
    ws[f"A{ROW}"].font = FONT_SECTION
    ws[f"A{ROW}"].fill = SUBHEADER
    ROW += 1

    waste_headers = ["Kategori", "Total (kg)", "% dari Total Waste"]
    for c, h in enumerate(waste_headers, 1):
        ws.cell(ROW, c, h).font = FONT_HEADER
        ws.cell(ROW, c).fill = HEADER
        ws.cell(ROW, c).alignment = CENTER
        ws.cell(ROW, c).border = THIN
    ROW += 1

    total_waste_nasi = round(sum(d.get("waste_nasi", 0) for d in report["days"]), 2)
    total_waste_sayur = round(sum(d.get("waste_sayur", 0) for d in report["days"]), 2)
    total_waste_lauk = round(sum(d.get("waste_lauk", 0) for d in report["days"]), 2)
    total_waste_buah = round(sum(d.get("waste_buah", 0) for d in report["days"]), 2)
    all_waste = total_waste_nasi + total_waste_sayur + total_waste_lauk + total_waste_buah

    waste_rows = [
        ("🍚 Karbohidrat (Nasi, Kentang, dll)", total_waste_nasi),
        ("🥗 Sayur", total_waste_sayur),
        ("🍗 Protein (Lauk)", total_waste_lauk),
        ("🍈 Buah", total_waste_buah),
    ]

    for label, val in waste_rows:
        ws.cell(ROW, 1, label).font = FONT_BODY
        ws.cell(ROW, 2, f"{val} kg").font = FONT_BODY
        ws.cell(ROW, 3, f"{(val/all_waste*100) if all_waste else 0:.1f}%").font = FONT_BODY
        for c in range(1, 4):
            ws.cell(ROW, c).border = THIN
            ws.cell(ROW, c).alignment = CENTER
        # Highlight biggest waste
        if val > 0 and val == max([w[1] for w in waste_rows]):
            ws.cell(ROW, 1).fill = WARN
            ws.cell(ROW, 2).fill = WARN
            ws.cell(ROW, 3).fill = WARN
        ROW += 1

    ws.cell(ROW, 1, "TOTAL WASTE").font = FONT_BOLD
    ws.cell(ROW, 2, f"{all_waste} kg").font = FONT_BOLD
    ws.cell(ROW, 3, "100%").font = FONT_BOLD
    for c in range(1, 4):
        ws.cell(ROW, c).border = THIN
        ws.cell(ROW, c).fill = SUBHEADER
        ws.cell(ROW, c).alignment = CENTER
    ROW += 2

    # ── SECTION 4: MASALAH OPERASIONAL ──
    ws.merge_cells(f"A{ROW}:H{ROW}")
    ws[f"A{ROW}"].value = "⚠️  REKAP MASALAH OPERASIONAL"
    ws[f"A{ROW}"].font = Font(name="Calibri", bold=True, size=11, color="C00000")
    ws[f"A{ROW}"].fill = RED_FILL
    ROW += 1

    masalah_headers = ["Tanggal", "Masalah", "Tindakan", "Teratasi"]
    for c, h in enumerate(masalah_headers, 1):
        ws.cell(ROW, c, h).font = FONT_HEADER
        ws.cell(ROW, c).fill = HEADER
        ws.cell(ROW, c).alignment = CENTER
        ws.cell(ROW, c).border = THIN
    ROW += 1

    ada_masalah = False
    for d in report["days"]:
        if d.get("masalah") and str(d["masalah"]).strip() and d["masalah"] != "-":
            ada_masalah = True
            ws.cell(ROW, 1, tgl_indo(d["tanggal"])).font = FONT_BODY
            ws.cell(ROW, 2, d["masalah"]).font = FONT_BODY
            ws.cell(ROW, 2).alignment = LEFT
            ws.cell(ROW, 3, d.get("tindakan", "")).font = FONT_BODY
            ws.cell(ROW, 3).alignment = LEFT
            ws.cell(ROW, 4, "✅" if d.get("teratasi") == "y" else "❌").font = FONT_BODY
            for c in range(1, 5):
                ws.cell(ROW, c).border = THIN
                ws.cell(ROW, c).alignment = CENTER if c != 2 and c != 3 else LEFT
            ROW += 1

    # Feedback PM section
    feedbacks = [d for d in report["days"] if d.get("feedback_pm") and str(d["feedback_pm"]).strip()]
    if feedbacks:
        ROW += 1
        ws.merge_cells(f"A{ROW}:H{ROW}")
        ws[f"A{ROW}"].value = "💬 FEEDBACK PENERIMA MANFAAT"
        ws[f"A{ROW}"].font = FONT_SECTION
        ws[f"A{ROW}"].fill = SUBHEADER
        ROW += 1
        for d in feedbacks:
            ws.cell(ROW, 1, tgl_indo(d["tanggal"])).font = FONT_BOLD
            ws.cell(ROW, 2, d["feedback_pm"]).font = FONT_BODY
            ws.merge_cells(f"B{ROW}:H{ROW}")
            for c in range(1, 9):
                ws.cell(ROW, c).border = THIN
                ws.cell(ROW, c).alignment = LEFT
            ROW += 1

    ROW += 1

    # ── SECTION 5: ANALISA ROOT CAUSE ──
    if report.get("analisa"):
        ws.merge_cells(f"A{ROW}:H{ROW}")
        ws[f"A{ROW}"].value = "📈 ROOT CAUSE ANALYSIS"
        ws[f"A{ROW}"].font = FONT_SECTION
        ws[f"A{ROW}"].fill = SUBHEADER
        ROW += 1
        ws.cell(ROW, 1, "Temuan").font = FONT_BOLD
        ws.cell(ROW, 2, report["analisa"].get("temuan", "")).font = FONT_BODY
        ws.merge_cells(f"B{ROW}:H{ROW}")
        for c in range(1, 9):
            ws.cell(ROW, c).border = THIN
            ws.cell(ROW, c).alignment = LEFT
        ROW += 1
        ws.cell(ROW, 1, "Root Cause").font = FONT_BOLD
        ws.cell(ROW, 2, report["analisa"].get("root_cause", "")).font = FONT_BODY
        ws.merge_cells(f"B{ROW}:H{ROW}")
        for c in range(1, 9):
            ws.cell(ROW, c).border = THIN
            ws.cell(ROW, c).alignment = LEFT
        ROW += 1
        ws.cell(ROW, 1, "Tindakan Korektif").font = FONT_BOLD
        ws.cell(ROW, 2, report["analisa"].get("tindakan_korektif", "")).font = FONT_BODY
        ws.merge_cells(f"B{ROW}:H{ROW}")
        for c in range(1, 9):
            ws.cell(ROW, c).border = THIN
            ws.cell(ROW, c).alignment = LEFT
        ROW += 2

    # ── SECTION 6: EVALUASI PER DIVISI ──
    if report.get("evaluasi_divisi"):
        ws.merge_cells(f"A{ROW}:H{ROW}")
        ws[f"A{ROW}"].value = "📌 EVALUASI PER DIVISI"
        ws[f"A{ROW}"].font = FONT_SECTION
        ws[f"A{ROW}"].fill = SUBHEADER
        ROW += 1

        ev_headers = ["Divisi", "Status", "Temuan", "Rekomendasi"]
        for c, h in enumerate(ev_headers, 1):
            ws.cell(ROW, c, h).font = FONT_HEADER
            ws.cell(ROW, c).fill = HEADER
            ws.cell(ROW, c).alignment = CENTER
            ws.cell(ROW, c).border = THIN
        ROW += 1

        for d in report["evaluasi_divisi"]:
            status = d.get("status", "")
            ws.cell(ROW, 1, d.get("divisi", "")).font = FONT_BOLD
            ws.cell(ROW, 2, f"🟢 {status}" if status == "Sangat Baik" else f"🟡 {status}" if "Evaluasi" in status else f"🔴 {status}").font = FONT_BODY
            ws.cell(ROW, 3, d.get("temuan", "")).font = FONT_BODY
            ws.cell(ROW, 4, d.get("rekomendasi", "")).font = FONT_BODY
            for c in range(1, 5):
                ws.cell(ROW, c).border = THIN
                ws.cell(ROW, c).alignment = LEFT
                ws.cell(ROW, c).fill = GREEN_FILL if status == "Sangat Baik" else YELLOW_FILL if "Evaluasi" in status else RED_FILL
            ROW += 1
        ROW += 1

    # ── SECTION 7: REKOMENDASI ──
    if report.get("rekomendasi"):
        ws.merge_cells(f"A{ROW}:H{ROW}")
        ws[f"A{ROW}"].value = "🎯 REKOMENDASI"
        ws[f"A{ROW}"].font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
        ws[f"A{ROW}"].fill = GREEN_FILL
        ROW += 1
        for i, rec in enumerate(report["rekomendasi"], 1):
            ws.cell(ROW, 1, f"{i}.").font = FONT_BOLD
            ws.cell(ROW, 2, rec).font = FONT_BODY
            ws.merge_cells(f"B{ROW}:H{ROW}")
            for c in range(1, 9):
                ws.cell(ROW, c).border = THIN
                ws.cell(ROW, c).alignment = LEFT
            ROW += 1
        ROW += 1

    # ── SECTION 8: EXECUTIVE SUMMARY ──
    if report.get("executive_summary"):
        ws.merge_cells(f"A{ROW}:H{ROW}")
        ws[f"A{ROW}"].value = "🔥 EXECUTIVE SUMMARY"
        ws[f"A{ROW}"].font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
        ws[f"A{ROW}"].fill = SUBHEADER
        ROW += 1
        ws.cell(ROW, 1, report["executive_summary"]).font = FONT_BODY
        ws.merge_cells(f"A{ROW}:H{ROW}")
        ws.cell(ROW, 1).alignment = LEFT
        for c in range(1, 9):
            ws.cell(ROW, c).border = THIN
        ROW += 2

    # ── SECTION 9: INSIGHT & REKOMENDASI ──
    if report.get("insight"):
        ws.merge_cells(f"A{ROW}:H{ROW}")
        ws[f"A{ROW}"].value = "💡 INSIGHT & INPUTAN"
        ws[f"A{ROW}"].font = FONT_SECTION
        ws[f"A{ROW}"].fill = SUBHEADER
        ROW += 1
        ws.cell(ROW, 1, report["insight"]).font = FONT_BODY
        ws.merge_cells(f"A{ROW}:H{ROW}")
        ws.cell(ROW, 1).alignment = LEFT
        for c in range(1, 9):
            ws.cell(ROW, c).border = THIN
        ROW += 2

    # Signature
    ws.cell(ROW, 1, f"Disusun oleh: {report.get('penyusun', 'Asisten Lapangan')}").font = Font(name="Calibri", italic=True, size=9, color="999999")
    ws.cell(ROW + 1, 1, f"Tgl laporan: {tgl_indo(report.get('tgl_laporan', str(date.today())))}").font = Font(name="Calibri", italic=True, size=9, color="999999")

    # Column widths
    widths = [5, 18, 15, 15, 15, 14, 14, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(w, 10)

    wb.save(filepath)
    return filepath


def input_wizard():
    """Interactive CLI wizard for weekly report"""
    print("\n" + "="*50)
    print("📊 LAPORAN MINGGUAN SPPG")
    print("="*50)

    report = {}
    report["nama_sppg"] = input("\n🏢 Nama SPPG [SPPG Wonodri 3]: ").strip() or "SPPG Wonodri 3"
    report["penyusun"] = input("👤 Disusun oleh [Asisten Lapangan]: ").strip() or "Asisten Lapangan"

    # Date range
    print("\n📅 Periode laporan mingguan:")
    tgl_start = input("  Tanggal mulai (YYYY-MM-DD): ").strip()
    tgl_end = input("  Tanggal selesai (YYYY-MM-DD): ").strip()

    days, start_d, end_d = load_week_data(tgl_start if tgl_start else None, tgl_end if tgl_end else None)
    if not tgl_start:
        print(f"\n  → Auto: {tgl_indo(str(start_d))} – {tgl_indo(str(end_d))}")
    report["tgl_start"] = str(start_d)
    report["tgl_end"] = str(end_d)
    report["tgl_laporan"] = str(date.today())

    print(f"\n📂 Data harian ditemukan: {len([d for d in days if d['target'] > 0])} hari")
    for d in days:
        if d["target"] > 0:
            print(f"  • {tgl_indo(d['tanggal'])}: {d['target']} target, {d['produksi']} produksi")

    report["days"] = days

    # ── Analisa ──
    print("\n" + "-"*50)
    print("📈 ROOT CAUSE ANALYSIS")
    print("-"*50)
    report["analisa"] = {}
    report["analisa"]["temuan"] = input("Temuan utama minggu ini: ").strip()
    report["analisa"]["root_cause"] = input("Root cause: ").strip()
    report["analisa"]["tindakan_korektif"] = input("Tindakan korektif: ").strip()

    # ── Evaluasi Divisi ──
    print("\n" + "-"*50)
    print("📌 EVALUASI PER DIVISI")
    print("-"*50)
    divisi_list = ["Admin Gudang", "Persiapan", "Cooking", "Pemorsian", "Distribusi", "Cuci Ompreng"]
    report["evaluasi_divisi"] = []
    for div in divisi_list:
        print(f"\n  {div}:")
        status = input("    Status (Sangat Baik/Perlu Evaluasi/Kurang): ").strip() or "Sangat Baik"
        temuan = input("    Temuan (jika ada): ").strip()
        rekom = input("    Rekomendasi: ").strip()
        report["evaluasi_divisi"].append({
            "divisi": div,
            "status": status,
            "temuan": temuan,
            "rekomendasi": rekom
        })

    # ── Rekomendasi ──
    print("\n" + "-"*50)
    print("🎯 REKOMENDASI (enter 2x untuk selesai)")
    print("-"*50)
    report["rekomendasi"] = []
    i = 1
    while True:
        r = input(f"  {i}. ").strip()
        if not r:
            break
        report["rekomendasi"].append(r)
        i += 1

    # ── Executive Summary ──
    print("\n" + "-"*50)
    print("🔥 EXECUTIVE SUMMARY")
    print("-"*50)
    report["executive_summary"] = input("  Executive summary: ").strip()

    # ── Insight ──
    print("\n" + "-"*50)
    print("💡 INSIGHT & INPUTAN")
    print("-"*50)
    report["insight"] = input("  Insight / catatan tambahan: ").strip()

    # ── Status ──
    print("\n" + "-"*50)
    print("🏆 STATUS PERFORMA")
    print("-"*50)
    report["status_performa"] = input("  Status (Baik/Cukup/Kurang) [Baik]: ").strip() or "Baik"

    return report


def main():
    init_dirs()
    report = input_wizard()

    # Save JSON
    tgl_key = f"{report['tgl_start']}_to_{report['tgl_end']}"
    json_path = MINGGUAN_DIR / f"mingguan_{tgl_key}.json"
    save_json(json_path, report)

    # Save Excel
    xlsx_path = MINGGUAN_DIR / f"laporan_mingguan_{tgl_key}.xlsx"
    render_excel(report, xlsx_path)

    # ── Print Ringkasan ──
    total_prod = sum(d["produksi"] for d in report["days"])
    print("\n" + "="*50)
    print("✅ LAPORAN MINGGUAN TERSIMPAN")
    print(f"📄 Excel: {xlsx_path}")
    print(f"📋 JSON:  {json_path}")
    print(f"\n📊 Ringkasan:")
    print(f"  Total Produksi: {total_prod:,} porsi")
    print(f"  Periode: {tgl_indo(report['tgl_start'])} – {tgl_indo(report['tgl_end'])}")
    print("="*50)


if __name__ == "__main__":
    main()
