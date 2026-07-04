#!/usr/bin/env python3
"""
LAPORAN HARIAN SPPG — Daily Operational Report
Mode: CLI input wizard + export ke Excel + JSON
"""

import sys
import json
from pathlib import Path
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Add tools dir to path
sys.path.insert(0, str(Path(__file__).parent))
from utils import (
    DATA_DIR, HARIAN_DIR, TEMPLATE_DIR,
    tgl_str, tgl_indo, nama_hari, rupiah, save_json, load_json, init_dirs
)

# ── Styling ──
HEADER = PatternFill("solid", fgColor="1F4E79")
SUBHEADER = PatternFill("solid", fgColor="D6E4F0")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
WARN = PatternFill("solid", fgColor="FFF2CC")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")

FONT_TITLE = Font(name="Calibri", bold=True, size=14, color="1F4E79")
FONT_HEADER = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", bold=True, size=10)
FONT_BODY = Font(name="Calibri", size=10)

THIN = Border(
    left=Side("thin", "B4C6E7"), right=Side("thin", "B4C6E7"),
    top=Side("thin", "B4C6E7"), bottom=Side("thin", "B4C6E7"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def input_wizard():
    """Interactive CLI input wizard"""
    print("\n" + "="*50)
    print("📋 LAPORAN OPERASIONAL HARIAN SPPG")
    print("="*50)

    data = {}

    # Tanggal & identitas
    tgl_input = input(f"\n📅 Tanggal [{date.today().strftime('%Y-%m-%d')}]: ").strip()
    data["tanggal"] = tgl_input if tgl_input else str(date.today())
    data["hari"] = nama_hari(data["tanggal"])
    data["nama_sppg"] = input("🏢 Nama SPPG [SPPG Wonodri 3]: ").strip() or "SPPG Wonodri 3"
    data["id_sppg"] = input("🆔 ID SPPG: ").strip()
    data["penyusun"] = input("👤 Disusun oleh [Asisten Lapangan]: ").strip() or "Asisten Lapangan"

    print("\n" + "-"*50)
    print("📦 PRODUKSI & DISTRIBUSI")
    print("-"*50)

    data["target"] = int(input("🎯 Target porsi hari ini: ") or 0)
    data["produksi"] = int(input("🍳 Produksi: ") or 0)
    data["distribusi"] = int(input("🚚 Distribusi: ") or 0)
    tepat = input("✅ Tepat waktu? (y/n) [y]: ").strip().lower() or "y"
    data["tepat_waktu"] = tepat == "y"
    kebersihan = input("🧹 Kebersihan (baik/cukup/kurang) [baik]: ").strip() or "baik"
    data["kebersihan"] = kebersihan

    # Breakdown per kategori
    print("\n" + "-"*50)
    print("👥 RINCIAN PENERIMA MANFAAT")
    print("-"*50)
    print("Kelompok BESAR (SD kls 4-6, SMP, SMA, ATS 9-18): Rp10.000")
    data["jml_besar"] = int(input("  Jumlah: ") or 0)
    print("Kelompok TENDIK (Pendidik + Tenaga Kependidikan):")
    data["jml_tendik"] = int(input("  Jumlah: ") or 0)
    print("Kelompok KECIL (PAUD/TK, SD kls 1-3, Balita, ATS <9): Rp8.000")
    data["jml_kecil"] = int(input("  Jumlah: ") or 0)
    print("Kelompok 3B (Ibu Hamil, Menyusui, Balita 6-59bln):")
    data["jml_3b"] = int(input("  Jumlah: ") or 0)
    data["total_pm"] = data["jml_besar"] + data["jml_tendik"] + data["jml_kecil"] + data["jml_3b"]

    # Menu
    print("\n" + "-"*50)
    print("🍽️  MENU HARI INI")
    print("-"*50)
    data["menu_karbohidrat"] = input("Karbohidrat [Nasi Putih]: ").strip() or "Nasi Putih"
    data["menu_protein_hewani"] = input("Protein Hewani [Ayam]: ").strip() or "Ayam"
    data["menu_protein_nabati"] = input("Protein Nabati [Tahu]: ").strip() or "Tahu"
    data["menu_sayur"] = input("Sayur [Sayur Kol Wortel]: ").strip() or "Sayur Kol Wortel"
    data["menu_buah"] = input("Buah [Jeruk]: ").strip() or "Jeruk"

    # Waste
    print("\n" + "-"*50)
    print("🗑️  WASTE MAKANAN (kg)")
    print("-"*50)
    data["waste_nasi"] = float(input("  Nasi (kg) [0]: ") or 0)
    data["waste_sayur"] = float(input("  Sayur (kg) [0]: ") or 0)
    data["waste_lauk"] = float(input("  Lauk (kg) [0]: ") or 0)
    data["waste_buah"] = float(input("  Buah (kg) [0]: ") or 0)
    data["waste_total"] = round(data["waste_nasi"] + data["waste_sayur"] + data["waste_lauk"] + data["waste_buah"], 2)

    # Masalah
    print("\n" + "-"*50)
    print("⚠️  MASALAH OPERASIONAL")
    print("-"*50)
    masalah_ada = input("Ada masalah? (y/n) [n]: ").strip().lower() or "n"
    if masalah_ada == "y":
        data["masalah"] = input("  Deskripsi masalah: ").strip()
        data["tindakan"] = input("  Tindakan yang diambil: ").strip()
        data["kategori_masalah"] = input("  Kategori (near_miss/komplain/teknis/lain): ").strip()
        data["teratasi"] = input("✅ Teratasi? (y/n) [y]: ").strip().lower() or "y"
    else:
        data["masalah"] = ""
        data["tindakan"] = ""
        data["kategori_masalah"] = ""
        data["teratasi"] = "y"

    # Feedback PM
    print("\n" + "-"*50)
    print("💬 FEEDBACK PENERIMA MANFAAT (opsional)")
    print("-"*50)
    data["feedback_pm"] = input("  Feedback: ").strip()

    # Catatan tambahan
    print("\n" + "-"*50)
    print("📝 CATATAN")
    print("-"*50)
    data["catatan"] = input("  Catatan tambahan: ").strip()

    return data


def to_excel(data, filepath):
    """Generate Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Harian"

    # ── Title ──
    ws.merge_cells("A1:F1")
    ws["A1"].value = f"LAPORAN OPERASIONAL HARIAN SPPG — {tgl_indo(data['tanggal'])}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:F2")
    ws["A2"].value = data["nama_sppg"] + f" | {data['hari']}"
    ws["A2"].font = Font(name="Calibri", size=11, color="555555")
    ws["A2"].alignment = CENTER

    ROW = 4
    sections = [
        ("📦 PRODUKSI & DISTRIBUSI", [
            ("Target", str(data["target"]) + " porsi"),
            ("Produksi", str(data["produksi"]) + " porsi"),
            ("Distribusi", str(data["distribusi"]) + " porsi"),
            ("Tepat Waktu", "✅ Ya" if data["tepat_waktu"] else "❌ Tidak"),
            ("Kebersihan", data["kebersihan"].title()),
            ("% Pencapaian", f"{(data['produksi']/data['target']*100) if data['target'] else 0:.0f}%"),
        ]),
        ("👥 PENERIMA MANFAAT", [
            ("Kelompok Besar (Rp10.000)", str(data["jml_besar"])),
            ("Tenaga Kependidikan", str(data["jml_tendik"])),
            ("Kelompok Kecil (Rp8.000)", str(data["jml_kecil"])),
            ("Kelompok 3B", str(data["jml_3b"])),
            ("TOTAL Penerima Manfaat", str(data["total_pm"])),
        ]),
        ("🍽️  MENU HARI INI", [
            ("Karbohidrat", data["menu_karbohidrat"]),
            ("Protein Hewani", data["menu_protein_hewani"]),
            ("Protein Nabati", data["menu_protein_nabati"]),
            ("Sayur", data["menu_sayur"]),
            ("Buah", data["menu_buah"]),
        ]),
        ("🗑️  WASTE MAKANAN (kg)", [
            ("Nasi", f"{data['waste_nasi']} kg"),
            ("Sayur", f"{data['waste_sayur']} kg"),
            ("Lauk", f"{data['waste_lauk']} kg"),
            ("Buah", f"{data['waste_buah']} kg"),
            ("TOTAL WASTE", f"{data['waste_total']} kg"),
        ]),
    ]

    for title, items in sections:
        # Section title
        ws.merge_cells(f"A{ROW}:F{ROW}")
        ws[f"A{ROW}"].value = title
        ws[f"A{ROW}"].font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
        ws[f"A{ROW}"].fill = SUBHEADER
        ws[f"A{ROW}"].alignment = LEFT
        ROW += 1

        for label, value in items:
            ws[f"A{ROW}"].value = label
            ws[f"A{ROW}"].font = FONT_BOLD
            ws[f"A{ROW}"].fill = LIGHT_GRAY
            ws[f"B{ROW}"].value = value
            ws[f"B{ROW}"].font = FONT_BODY
            ws[f"C{ROW}"].value = ""
            for c in range(1, 4):
                ws.cell(ROW, c).border = THIN
                ws.cell(ROW, c).alignment = LEFT
            ROW += 1
        ROW += 1  # space

    # ── Masalah ──
    ws.merge_cells(f"A{ROW}:F{ROW}")
    ws[f"A{ROW}"].value = "⚠️  MASALAH OPERASIONAL"
    ws[f"A{ROW}"].font = Font(name="Calibri", bold=True, size=11, color="C00000")
    ws[f"A{ROW}"].fill = WARN
    ROW += 1

    if data["masalah"]:
        for label, val in [("Deskripsi", data["masalah"]), ("Tindakan", data["tindakan"]),
                           ("Kategori", data["kategori_masalah"]),
                           ("Teratasi", "✅ Ya" if data["teratasi"] else "❌ Belum")]:
            ws[f"A{ROW}"].value = label
            ws[f"A{ROW}"].font = FONT_BOLD
            ws[f"B{ROW}"].value = val
            ws[f"B{ROW}"].font = FONT_BODY
            ws.merge_cells(f"B{ROW}:D{ROW}")
            for c in range(1, 5):
                ws.cell(ROW, c).border = THIN
                ws.cell(ROW, c).alignment = LEFT
            ROW += 1
    else:
        ws[f"A{ROW}"].value = "✅ Tidak ditemukan masalah operasional"
        ws[f"A{ROW}"].font = FONT_BODY
        ws[f"A{ROW}"].fill = GREEN_FILL
        ROW += 1

    ROW += 1
    # ── Feedback ──
    ws.merge_cells(f"A{ROW}:F{ROW}")
    ws[f"A{ROW}"].value = "💬 FEEDBACK PENERIMA MANFAAT"
    ws[f"A{ROW}"].font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    ws[f"A{ROW}"].fill = SUBHEADER
    ROW += 1
    ws[f"A{ROW}"].value = data["feedback_pm"] if data["feedback_pm"] else "—"
    ws[f"A{ROW}"].font = FONT_BODY
    ws[f"A{ROW}"].alignment = LEFT
    ROW += 2

    # ── Catatan ──
    ws.merge_cells(f"A{ROW}:F{ROW}")
    ws[f"A{ROW}"].value = "📝 CATATAN"
    ws[f"A{ROW}"].font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    ws[f"A{ROW}"].fill = SUBHEADER
    ROW += 1
    ws[f"A{ROW}"].value = data["catatan"] if data["catatan"] else "—"
    ws[f"A{ROW}"].font = FONT_BODY
    ws[f"A{ROW}"].alignment = LEFT
    ROW += 2

    # ── Signature ──
    ws.merge_cells(f"A{ROW}:F{ROW}")
    ws[f"A{ROW}"].value = f"Disusun oleh: {data['penyusun']} | WIB, {tgl_indo(data['tanggal'])}"
    ws[f"A{ROW}"].font = Font(name="Calibri", italic=True, size=9, color="999999")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    wb.save(filepath)
    return filepath


def main():
    init_dirs()
    data = input_wizard()

    # Save JSON
    tgl = data["tanggal"]
    json_path = HARIAN_DIR / f"{tgl}.json"
    save_json(json_path, data)

    # Save Excel
    xlsx_path = HARIAN_DIR / f"laporan_harian_{tgl}.xlsx"
    to_excel(data, xlsx_path)

    print("\n" + "="*50)
    print(f"✅ LAPORAN HARIAN TERSIMPAN")
    print(f"📄 Excel: {xlsx_path}")
    print(f"📋 JSON:  {json_path}")
    print("="*50)


if __name__ == "__main__":
    main()
