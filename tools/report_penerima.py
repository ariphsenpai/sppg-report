#!/usr/bin/env python3
"""
REPORT PENERIMA MANFAAT — Beneficiary Update Tracker
Data source: Google Sheets CSV export or manual input
"""

import sys
import json
from pathlib import Path
from datetime import date, datetime, timedelta
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.request import urlopen

sys.path.insert(0, str(Path(__file__).parent))
from utils import (
    DATA_DIR, PENERIMA_DIR, TEMPLATE_DIR,
    tgl_str, tgl_indo, rupiah, save_json, load_json, init_dirs
)

# ── Styling ──
HEADER = PatternFill("solid", fgColor="1F4E79")
SUBHEADER = PatternFill("solid", fgColor="D6E4F0")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")

FONT_TITLE = Font(name="Calibri", bold=True, size=14, color="1F4E79")
FONT_HEADER = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", bold=True, size=10)
FONT_BODY = Font(name="Calibri", size=10)

THIN = Border(
    left=Side("thin", "B4C6E7"), right=Side("thin", "B4C6E7"),
    top=Side("thin", "B4C6E7"), bottom=Side("thin", "B4C6E7"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def parse_sheet_url(url):
    """Convert Google Sheets URL to CSV export URL"""
    # Already a CSV export URL?
    if "export?format=csv" in url or "export?format=json" in url:
        return url
    if "docs.google.com/spreadsheets" in url:
        # Extract ID
        if "/d/" in url:
            sheet_id = url.split("/d/")[1].split("/")[0]
        elif "id=" in url:
            sheet_id = url.split("id=")[1].split("&")[0]
        else:
            return None
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
    return url


def fetch_from_gsheet(url):
    """Fetch CSV data from Google Sheets"""
    csv_url = parse_sheet_url(url)
    if not csv_url:
        print("❌ URL Google Sheets tidak valid")
        return None
    try:
        resp = urlopen(csv_url, timeout=15)
        data = resp.read().decode("utf-8")
        reader = csv.DictReader(io.StringIO(data))
        rows = list(reader)
        print(f"✅ Berhasil fetch {len(rows)} baris data")
        return rows
    except Exception as e:
        print(f"❌ Gagal fetch data: {e}")
        return None


def fetch_from_csv(filepath):
    """Load from local CSV file"""
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader)


def render_excel(rows, source_desc, filepath, nama_sppg="SPPG Wonodri 3"):
    """Generate Excel report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Penerima Manfaat"

    ROW = 1
    ws.merge_cells("A1:H1")
    ws["A1"].value = f"📋 DATA PENERIMA MANFAAT — {nama_sppg}"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = CENTER
    ROW = 2
    ws.merge_cells(f"A{ROW}:H{ROW}")
    ws[f"A{ROW}"].value = f"Sumber: {source_desc} | Tgl update: {tgl_indo(str(date.today()))}"
    ws[f"A{ROW}"].font = Font(name="Calibri", size=11, color="555555")
    ws[f"A{ROW}"].alignment = CENTER
    ROW = 4

    # ── HEADERS ──
    fieldnames = rows[0].keys() if rows else []
    headers_map = {
        "Waktu": "Waktu Distribusi",
        "Tempat": "Satuan Pendidikan/Posyandu",
        "Besar": "Besar (SD 4-6/SMP/SMA/ATS 9-18)",
        "Tendik": "Tendik (Pendidik & Tenaga Kependidikan)",
        "Kecil": "Kecil (PAUD/TK/SD 1-3/Balita/ATS<9)",
        "Total": "Total",
    }
    headers_display = []
    for fn in fieldnames:
        headers_display.append(headers_map.get(fn, fn))

    # Sheet 1: Detail Penerima
    for c, h in enumerate(headers_display, 1):
        cell = ws.cell(ROW, c, h)
        cell.font = FONT_HEADER
        cell.fill = HEADER
        cell.alignment = CENTER
        cell.border = THIN
    ROW += 1

    total_all = 0
    total_besar = 0
    total_tendik = 0
    total_kecil = 0

    for i, row in enumerate(rows):
        r = ROW + i
        for c, fn in enumerate(fieldnames, 1):
            val = row.get(fn, "")
            cell = ws.cell(r, c, val)
            cell.font = FONT_BODY
            cell.border = THIN
            cell.alignment = CENTER if fn != "Tempat" else LEFT
            # Count totals
            if fn == "Besar":
                try:
                    total_besar += int(val) if val else 0
                except:
                    pass
            elif fn == "Tendik":
                try:
                    total_tendik += int(val) if val else 0
                except:
                    pass
            elif fn == "Kecil":
                try:
                    total_kecil += int(val) if val else 0
                except:
                    pass
            elif fn == "Total":
                try:
                    total_all += int(val) if val else 0
                except:
                    pass

        if (i % 2) == 1:
            for c in range(1, len(fieldnames) + 1):
                ws.cell(r, c).fill = LIGHT_GRAY

    # Grand total row
    total_row = ROW + len(rows)
    if "Tempat" in fieldnames:
        ws.cell(total_row, 1, "").font = FONT_BOLD
        ws.cell(total_row, 2, "GRAND TOTAL").font = FONT_BOLD
    else:
        ws.cell(total_row, 1, "GRAND TOTAL").font = FONT_BOLD

    col_map = {}
    for i, fn in enumerate(fieldnames, 1):
        col_map[fn] = i
    if "Besar" in col_map:
        ws.cell(total_row, col_map["Besar"], total_besar).font = FONT_BOLD
    if "Tendik" in col_map:
        ws.cell(total_row, col_map["Tendik"], total_tendik).font = FONT_BOLD
    if "Kecil" in col_map:
        ws.cell(total_row, col_map["Kecil"], total_kecil).font = FONT_BOLD
    if "Total" in col_map:
        ws.cell(total_row, col_map["Total"], total_all).font = FONT_BOLD

    for c in range(1, len(fieldnames) + 1):
        ws.cell(total_row, c).border = THIN
        ws.cell(total_row, c).fill = SUBHEADER
        ws.cell(total_row, c).alignment = CENTER

    # ── Auto width ──
    for c in range(1, len(fieldnames) + 1):
        max_len = max(
            len(str(ws.cell(r, c).value or ""))
            for r in range(1, total_row + 1)
        )
        ws.column_dimensions[get_column_letter(c)].width = max(min(max_len + 3, 50), 12)

    # ── Sheet 2: Summary ──
    ws2 = wb.create_sheet("Rekap")
    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "REKAP PENERIMA MANFAAT"
    ws2["A1"].font = FONT_TITLE
    ws2["A1"].alignment = CENTER
    ROW2 = 3

    summary_headers = ["Kategori", "Total", "% dari Total", "Estimasi Biaya/Hari"]
    for c, h in enumerate(summary_headers, 1):
        cell = ws2.cell(ROW2, c, h)
        cell.font = FONT_HEADER
        cell.fill = HEADER
        cell.alignment = CENTER
        cell.border = THIN
    ROW2 += 1

    biaya_besar = total_besar * 10000
    biaya_kecil = total_kecil * 8000
    biaya_tendik = total_tendik * 0  # Tenaga kependidikan free?
    # Actually tendik may have its own cost structure - assume same as besar
    biaya_tendik = total_tendik * 10000
    total_biaya = biaya_besar + biaya_kecil + biaya_tendik

    summaries = [
        ("Besar (SD 4-6/SMP/SMA/ATS 9-18)", total_besar, biaya_besar),
        ("Tendik (Pendidik & Tendik)", total_tendik, biaya_tendik),
        ("Kecil (PAUD/TK/SD 1-3/Balita/ATS<9)", total_kecil, biaya_kecil),
    ]

    for label, jml, biaya in summaries:
        ws2.cell(ROW2, 1, label).font = FONT_BODY
        ws2.cell(ROW2, 2, jml).font = FONT_BODY
        ws2.cell(ROW2, 3, f"{(jml/total_all*100) if total_all else 0:.1f}%").font = FONT_BODY
        ws2.cell(ROW2, 4, rupiah(biaya)).font = FONT_BODY
        for c in range(1, 5):
            ws2.cell(ROW2, c).border = THIN
            ws2.cell(ROW2, c).alignment = CENTER
        ROW2 += 1

    # Total row
    ws2.cell(ROW2, 1, "TOTAL").font = FONT_BOLD
    ws2.cell(ROW2, 2, total_all).font = FONT_BOLD
    ws2.cell(ROW2, 3, "100%").font = FONT_BOLD
    ws2.cell(ROW2, 4, rupiah(total_biaya)).font = FONT_BOLD
    for c in range(1, 5):
        ws2.cell(ROW2, c).border = THIN
        ws2.cell(ROW2, c).fill = SUBHEADER
        ws2.cell(ROW2, c).alignment = CENTER

    # Institution breakdown
    ROW2 += 2
    ws2.cell(ROW2, 1, "RINCIAN PER INSTITUSI").font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    ws2.cell(ROW2, 1).fill = SUBHEADER
    ws2.merge_cells(f"A{ROW2}:D{ROW2}")
    ROW2 += 1

    ib_headers = ["Institusi", "Total PM", "Waktu Distribusi"]
    for c, h in enumerate(ib_headers, 1):
        cell = ws2.cell(ROW2, c, h)
        cell.font = FONT_HEADER
        cell.fill = HEADER
        cell.alignment = CENTER
        cell.border = THIN
    ROW2 += 1

    for row in rows:
        tempat = row.get("Tempat", "")
        total = row.get("Total", "")
        waktu = row.get("Waktu", "")
        ws2.cell(ROW2, 1, tempat).font = FONT_BODY
        ws2.cell(ROW2, 2, int(total) if total else 0).font = FONT_BODY
        ws2.cell(ROW2, 3, waktu).font = FONT_BODY
        for c in range(1, 4):
            ws2.cell(ROW2, c).border = THIN
            ws2.cell(ROW2, c).alignment = CENTER
        ROW2 += 1

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 22

    wb.save(filepath)
    return filepath


def main():
    init_dirs()

    print("\n" + "="*50)
    print("📋 REPORT PENERIMA MANFAAT SPPG")
    print("="*50)
    nama_sppg = input("\n🏢 Nama SPPG [SPPG Wonodri 3]: ").strip() or "SPPG Wonodri 3"

    print("\n📥 SUMBER DATA:")
    print("  1. Google Sheets URL")
    print("  2. File CSV lokal")
    print("  3. Input manual")
    choice = input("\nPilih sumber [1]: ").strip() or "1"

    rows = None
    source_desc = ""

    if choice == "1":
        url = input("\nGoogle Sheets URL: ").strip()
        if not url:
            print("❌ URL tidak boleh kosong")
            return
        rows = fetch_from_gsheet(url)
        if rows:
            source_desc = "Google Sheets"
    elif choice == "2":
        csv_path = input("Path file CSV: ").strip()
        try:
            rows = fetch_from_csv(csv_path)
            source_desc = f"CSV: {csv_path}"
        except Exception as e:
            print(f"❌ Gagal baca CSV: {e}")
            return
    elif choice == "3":
        print("\nMasukkan data manual (ketik 'selesai' untuk selesai):")
        print("Format: Waktu,Tempat,Besar,Tendik,Kecil")
        rows = []
        while True:
            line = input("  > ").strip()
            if line.lower() == "selesai" or not line:
                break
            parts = [p.strip() for p in line.split(",")]
            if len(parts) >= 5:
                rows.append({
                    "Waktu": parts[0],
                    "Tempat": parts[1],
                    "Besar": parts[2],
                    "Tendik": parts[3],
                    "Kecil": parts[4],
                    "Total": str(int(parts[2]) + int(parts[3]) + int(parts[4]))
                })
        source_desc = "Input Manual"

    if not rows:
        print("❌ Tidak ada data diproses")
        return

    # Save raw data
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = PENERIMA_DIR / f"penerima_{timestamp}.json"
    save_json(json_path, rows)

    # Generate Excel
    xlsx_path = PENERIMA_DIR / f"report_penerima_{timestamp}.xlsx"
    render_excel(rows, source_desc, xlsx_path, nama_sppg)

    print("\n" + "="*50)
    print("✅ REPORT PENERIMA MANFAAT TERSIMPAN")
    print(f"📄 Excel: {xlsx_path}")
    print(f"📋 JSON:  {json_path}")
    total_pm = sum(int(r.get("Total", 0)) for r in rows if r.get("Total", ""))
    print(f"👥 Total Penerima Manfaat: {total_pm:,}")
    print("="*50)


if __name__ == "__main__":
    main()
