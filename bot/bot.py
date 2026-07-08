#!/usr/bin/env python3
"""
SPPG Bot Telegram — Asisten Laporan Operasional SPPG
Powered by python-telegram-bot v22.x
"""

import os
import sys
import json
import io
import csv
import asyncio
from pathlib import Path
from datetime import date, datetime, timedelta
from typing import Optional, Dict, Any

# Add tools to path
sys.path.insert(0, str(Path(__file__).parent.parent / "tools"))
from utils import (
    HARIAN_DIR, MINGGUAN_DIR, PENERIMA_DIR, TEMPLATE_DIR,
    tgl_indo, tgl_str, nama_hari, rupiah, save_json, load_json, init_dirs,
    range_minggu
)

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
# In ptb v22.x, FSInputFile = InputFile (renamed in later versions)
from telegram._files.inputfile import InputFile as FSInputFile
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ConversationHandler,
    ContextTypes
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── States for ConversationHandler ──
(
    STATE_HARIAN_TGL,
    STATE_HARIAN_TARGET,
    STATE_HARIAN_PRODUKSI,
    STATE_HARIAN_DISTRIBUSI,
    STATE_HARIAN_BESAR,
    STATE_HARIAN_TENDIK,
    STATE_HARIAN_KECIL,
    STATE_HARIAN_3B,
    STATE_HARIAN_MENU_NASTAR,
    STATE_HARIAN_MENU_PROHE,
    STATE_HARIAN_MENU_PRONA,
    STATE_HARIAN_MENU_SAYUR,
    STATE_HARIAN_MENU_BUAH,
    STATE_HARIAN_WASTE_NASAK,
    STATE_HARIAN_WASTE_SAYUR,
    STATE_HARIAN_WASTE_LAUK,
    STATE_HARIAN_WASTE_BUAH,
    STATE_HARIAN_MASALAH,
    STATE_HARIAN_TINDAKAN,
    STATE_HARIAN_FEEDBACK,
    STATE_HARIAN_CATATAN,
    # Weekly
    STATE_MINGGUAN_TGL_START,
    STATE_MINGGUAN_TGL_END,
    STATE_MINGGUAN_ANALISA_TEMUAN,
    STATE_MINGGUAN_ANALISA_ROOT,
    STATE_MINGGUAN_ANALISA_TINDAKAN,
    STATE_MINGGUAN_REKOMENDASI,
    STATE_MINGGUAN_EXEC_SUMMARY,
    STATE_MINGGUAN_INSIGHT,
    # PM
    STATE_PENERIMA_SOURCE,
    STATE_PENERIMA_URL,
    STATE_PENERIMA_ROW,
    # Menu
    STATE_JUMLAH_REKOMENDASI,
) = range(33)

# ── Styling ──
HEADER = PatternFill("solid", fgColor="1F4E79")
SUBHEADER = PatternFill("solid", fgColor="D6E4F0")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
WARN = PatternFill("solid", fgColor="FFF2CC")

FONT_TITLE = Font(name="Calibri", bold=True, size=14, color="1F4E79")
FONT_HEADER = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", bold=True, size=10)
FONT_BODY = Font(name="Calibri", size=10)

THIN = Border(
    left=Side("thin", "B4C6E7"), right=Side("thin", "B4C6E7"),
    top=Side("thin", "B4C6E7"), bottom=Side("thin", "B4C6E7"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ── User data sessions ──
# Context.user_data keeps conversation state

JUKNIS_MENU = """
📚 *JUKNIS SPPG MBG 2026 — Quick Reference*

🏢 *STRUKTUR ORGANISASI SPPG*
• Kepala SPPG — Approver VA, koordinator
• Pengawas Gizi — Menu, QC, edukasi gizi
• Pengawas Keuangan — Anggaran, petty cash
• Pengawas Sanitasi — Higiene & sanitasi
• Asisten Lapangan — Hubungan eksternal, QC lapangan
• Juru Masak — Produksi makanan
• Relawan (max 52 org) — Persiapan, olah, porsi, packing, distribusi

💰 *BIAYA ACUAN (AT COST)*
• Balita/TK/SD kls 1-3 / ATS <9: Rp8.000/porsi
• SD 4-6/SMP/SMA/Tendik/3B: Rp10.000/porsi
• Operasional: Rp3.000/porsi
• Insentif Fasilitas SPPG: Rp6.000.000/hari (lumpsum)

📅 *HARI OPERASIONAL*
• 313 hari/tahun (365 - 52 Minggu)
• Senin - Sabtu
• Termasuk hari libur nasional & cuti bersama

👥 *PENERIMA MANFAAT*
• Peserta Didik: PAUD/TK s/d SMA/SMK + SLB + Pesantren
• Tendik: Pendidik & Tenaga Kependidikan
• 3B: Ibu Hamil, Ibu Menyusui, Balita 6-59bln
• Radius: max 6 km atau 30 menit tempuh

📋 *LAPORAN WAJIB*
• Harian: Rekap porsi (30a) + Penggunaan dana (30b)
• 2 Mingguan: Laporan penggunaan dana (30c)
• Bulanan: Summary + anggaran + PM (30d)
• Buku: Neraca, Petty Cash, Bahan Pangan, Operasional (30e-30h)
• Uji Organoleptik: setiap distribusi (Lamp 22)

🔬 *UJI ORGANOLEPTIK*
• 2x: saat tiba & sebelum dikonsumsi
• Parameter: Rasa (1-5), Warna (1-5), Aroma (1-5), Tekstur (1-5)
• Form dibawa sopir

⚡ *KETENTUAN*
• Bahan baku prioritas dari UMKM/BUMDes/Koperasi lokal
• Survei harga mingguan (min 3 penyedia per komoditas)
• Makanan wajib dikonsumsi max 4 jam setelah dimasak
• Kapasitas SPPG: max 2.500 PM (3.000 dgn juru masak bersertifikat)

📦 *PRODUKSI & DISTRIBUSI*
• Tahap: Penerimaan → Persiapan → Pengolahan → Pemorsian → Packing → Distribusi
• QC di setiap tahap: sortir, timbang, uji rasa/kematangan
• Packing: ompreng + label sesuai lokasi satuan Pendidikan
• Distribusi: kendaraan box alumunium, sopir bawa form organoleptik
• Penerimaan di lokasi: BAST tertanda, konsumsi max 4 jam
• Pengembalian ompreng: dicuci & disterilkan di SPPG

📊 *MONITORING & EVALUASI*
• Monitoring harian: produksi, distribusi, waste, masalah
• Input data ke SIPGN setiap hari
• Evaluasi menu: berdasarkan waste & feedback penerima
• Monitoring petty cash + insentif relawan
• Evaluasi kinerja relawan per shift
• Laporan bulanan evaluasi + tindakan korektif KM (Kejadian Menonjol)
"""


# ═══════════════════════════════════════
#  MENU HANDLERS
# ═══════════════════════════════════════

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Welcome + main menu"""
    init_dirs()
    keyboard = [
        [InlineKeyboardButton("📋 Laporan Harian", callback_data="menu_harian")],
        [InlineKeyboardButton("📊 Laporan Mingguan", callback_data="menu_mingguan")],
        [InlineKeyboardButton("📑 Report Penerima Manfaat", callback_data="menu_penerima")],
        [InlineKeyboardButton("📚 Juknis Quick Ref", callback_data="menu_juknis")],
        [InlineKeyboardButton("📈 Status Hari Ini", callback_data="menu_status")],
    ]
    reply = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        "🍽️ *SPPG REPORT TOOLS*\n"
        "Asisten Laporan Operasional SPPG — MBG\n\n"
        "Pilih menu di bawah:",
        reply_markup=reply,
        parse_mode="Markdown"
    )
JUKNIS_STRUKTUR = """
🏢 *STRUKTUR ORGANISASI SPPG*

*1. KEPALA SPPG (Ka.SPPG)*
• Status: SPPI (Sarjana Penggerak Pembangunan Indonesia)
• Tugas:
  - Koordinator pengelola SPPG
  - Approver Virtual Account (VA)
  - Penanggung jawab operasional SPPG
  - Tanda tangan laporan & BAST
  - Monitoring keseluruhan program
• Jumlah: 1 orang per SPPG
• Dibawah: Koordinator SPPG Kecamatan / KPPG

*2. PENGAWAS GIZI*
• Tugas:
  - Menyusun menu sesuai standar gizi (AKG)
  - Quality Control bahan pangan & makanan
  - Edukasi gizi & keamanan pangan
  - Pantau kecukupan gizi per kelompok sasaran
  - Evaluasi menu berdasarkan waste & feedback
• Jumlah: 1 orang per SPPG

*3. PENGAWAS KEUANGAN*
• Tugas:
  - Merencanakan anggaran belanja harian
  - Pengelolaan keuangan & kas kecil (petty cash)
  - Manajemen logistik bahan baku
  - Input laporan ke harian/2 mingguan/bulanan
  - Atur jadwal & insentif relawan
  - Rekonsiliasi dana VA
• Jumlah: 1 orang per SPPG

*4. PENGAWAS SANITASI*
• Tugas:
  - Inspeksi kebersihan lingkungan SPPG (dalam & luar)
  - Higiene sanitasi peralatan masak
  - Higiene sanitasi tim relawan sesuai SOP
  - Sistem pembuangan limbah
  - Suplai air bersih sesuai standar
• Jumlah: 1 orang per 5 SPPG (mobile)

*5. ASISTEN LAPANGAN (ASLAP)* ← *POSISI ANDA*
• Tugas sesuai JUKNIS Pasal 4.9.2:
  - Pelaksana terdepan hubungan eksternal
  - Koordinasi dengan satuan Pendidikan & Posyandu
  - Pastikan proses produksi selesai tepat waktu
  - Pastikan ketersediaan bahan baku
  - Update data porsi produksi & distribusi harian
  - Jaga hubungan harmonis dgn satuan Pendidikan
  - Monitoring & evaluasi packaging, distribusi, kebersihan, pencuci alat
  - Atur pengeluaran operasional (bensin, gas, dll)
  - QC, pencatatan & penimbangan bahan baku masuk
• Jumlah: 1 orang per SPPG

*6. JURU MASAK*
• Status: Minimal bersertifikat (untuk kapasitas >3.000 PM)
• Tugas:
  - Memasak & mengolah bahan makanan
  - Memastikan rasa & kematangan sesuai standar
  - Mengawasi tim pengolahan
• Jumlah: 1 orang per SPPG

*7. KOORDINATOR TIM (dari relawan)*
• Koordinator Tim Persiapan Bahan — 4 orang
• Koordinator Tim Pengolahan — 10 orang
• Koordinator Tim Pemorsian — 9 orang
• Koordinator Tim Packing — 1 orang
• Koordinator Tim Distribusi — 4 orang
• Koordinator Tim Kebersihan — 2 orang
• Koordinator Tim Pencuci Alat — 14 orang
• Keamanan — 2 orang
• *Total relawan: max 52 orang per SPPG*

*8. STRUKTUR ESELON DI ATASNYA*
• Koordinator SPPG Regional (Kreg) — Tingkat Provinsi
• Koordinator SPPG Wilayah (Korwil) — Tingkat Kab/Kota
• Koordinator SPPG Kecamatan — Tingkat Kecamatan
• KPPG — Kantor Pelayanan Pemenuhan Gizi
• BGN — Badan Gizi Nasional Pusat
"""

JUKNIS_BIAYA = """
💰 *BIAYA ACUAN PROGRAM MBG (AT COST)*

*Biaya Bahan Baku per Porsi:*
• Balita, PAUD/TK/RA, SD/MI kls 1-3, ATS <9th
  → Rp8.000/porsi/hari
• SD/MI kls 4-6, SMP/MTs, SMA/MA/SMK
  ATS 9-18th, Pendidik, Tendik, Ibu Hamil, Ibu Menyusui
  → Rp10.000/porsi/hari

*Biaya Operasional:*
• Rp3.000/porsi/hari (at cost — bisa kurang/lebih)
  Meliputi: listrik, gas, air, insentif relawan, insentif
  kader, insentif satuan pendidikan, BPJS ketenagakerjaan,
  BBM, sewa kendaraan, internet, pulsa, ATK, dll

*Insentif Fasilitas SPPG:*
• Rp6.000.000/hari (lumpsum — tetap per hari,
  tidak tergantung jumlah porsi)
• Dibayar per 2 mingguan

*CARA HITUNG:*
Biaya Per Hari = (Jumlah PM Kecil × 8.000) +
  (Jumlah PM Besar + Tendik + 3B × 10.000) +
  (Total PM × 3.000 operasional) + 6.000.000

*Sumber Acuan Harga:*
• HET (Harga Eceran Tertinggi) daerah setempat
• Survei mingguan: bandingkan 3 penyedia minimal
  per komoditas, ambil harga terendah
• At Cost = harga riil, bisa kurang/lebih dari acuan
"""

JUKNIS_HARI = """
📅 *HARI OPERASIONAL SPPG 2026*

*Total: 313 hari/tahun*
Perhitungan: 365 hari - 52 hari Minggu = 313 hari

*Jadwal:*
• Senin - Sabtu (6 hari kerja/minggu)
• Termasuk hari libur nasional & cuti bersama
• Hanya libur hari Minggu

*Hari Pendistribusian MBG:*
• Hari kerja termasuk libur nasional & cuti bersama
• Tetap jalan meskipun hari libur sekolah
  (khusus untuk kelompok 3B yang tetap dilayani)

*Jam Distribusi:*
• Makan Pagi: 06.00 - 09.00 (kontribusi 20-25% AKG)
• Makan Siang: 11.00 - 14.00 (kontribusi 30-35% AKG)

*Per Bulan:*
• Jumlah hari operasional per bulan bervariasi
  (24-27 hari tergantung jumlah hari Minggu)
• Contoh: Januari=27, Februari=24, Maret=26, dll
"""

JUKNIS_PENERIMA = """
👥 *PENERIMA MANFAAT MBG*

*KELOMPOK PESERTA DIDIK:*
1. PAUD/TK/RA/TK LB
2. SD/MI/SD SLB Kelas 1-3
3. SD/MI/SD SLB Kelas 4-6
4. SMP/MTs/SMP LB/Pesantren Kelas 1-3
5. SMA/MA/SMK/SMA LB/Pesantren Kelas 4-6
6. ATS (Anak Tidak Sekolah) Usia < 9 tahun
7. ATS Usia 9-18 tahun
8. Pendidik (Guru, Tutor, Dll)
9. Tenaga Kependidikan (Staff Sekolah)

*KELOMPOK NON PESERTA DIDIK (3B):*
10. Balita (6-59 bulan)
11. Ibu Hamil (semua trimester)
12. Ibu Menyusui (0-12 bulan pasca melahirkan)

*KETENTUAN:*
• Radius: max 6 km dari SPPG
  atau max 30 menit waktu tempuh
• Makanan wajib dikonsumsi max 4 jam setelah dimasak
• Kapasitas SPPG: max 2.500 PM
  (bisa 3.000 jika juru masak bersertifikat)
*Desil 1 & 2*: prioritas
"""

JUKNIS_LAPORAN = """
📋 *LAPORAN WAJIB SPPG*

*HARIAN:*
• Rekapitulasi Perhitungan Porsi (Lamp 30a)
• Laporan Harian Penggunaan Dana (Lamp 30b)
• Uji Organoleptik (Lamp 22) — setiap distribusi
• BAST MBG (Berita Acara Serah Terima)
• Daftar hadir Ka.SPPG + Pengawas Gizi + Keuangan

*2 MINGGUAN:*
• Laporan 2 Mingguan Penggunaan Dana (Lamp 30c)
  Dikirim ke KPPG/Direktur Wilayah

*BULANAN:*
• Laporan Bulanan — Summary (Lamp 30d A)
• Laporan Bulanan — Anggaran (Lamp 30d B)
• Laporan Bulanan — Penerima Manfaat (Lamp 30d C)

*BUKU BANTU (Real-time):*
• Buku Neraca Besar (Lamp 30e)
• Buku Petty Cash (Lamp 30f) — max Rp500rb/trans,
  max Rp5jt/minggu
• Buku Bahan Pangan (Lamp 30g)
• Buku Dana Operasional (Lamp 30h)
• Buku Insentif Fasilitas (Lamp 30i)

*LAPORAN LAIN:*
• Laporan Pelaksanaan Kegiatan (Lamp 16) — per 2 minggu
• Pernyataan Tanggung Jawab (Lamp 30j)
• Kuitansi & Bukti Tanda Terima (Lamp 30k, 30m)
• Daftar Nominatif Insentif Relawan (Lamp 30l)
• Berita Acara Pengalihan Sisa Dana (Lamp 30n)
• Surat Perjanjian Kerjasama (Lamp 25)
  — antara Ka.SPPG dgn Penerima Manfaat
"""

JUKNIS_GIZI = """
🔬 *STANDAR GIZI & UJI ORGANOLEPTIK*

*STANDAR GIZI:*
Acuan: Permenkes No. 28 Tahun 2019 tentang AKG
• Makan Pagi: Kontribusi 20-25% AKG
• Makan Siang: Kontribusi 30-35% AKG

*KOMPONEN MENU SEIMBANG:*
• Karbohidrat: Beras, Jagung, Ubi, Kentang, Mie, Sagu
• Protein Hewani: Ayam, Ikan, Telur, Daging, Udang
• Protein Nabati: Tempe, Tahu, Kacang-kacangan
• Sayur: Bayam, Kangkung, Wortel, Brokoli, Buncis, dll
• Buah: Pisang, Apel, Jeruk, Semangka, Pepaya, Melon

*GLOSARIUM BAHAN PANGAN (Lamp 31):*
Karbohidrat (14 jenis), Protein Hewani (14 jenis),
Protein Nabati (7 jenis), Sayur (24 jenis),
Buah (9 jenis), Bahan Baku Lain (20 jenis)

*UJI ORGANOLEPTIK (Lamp 22):*
• Dilakukan 2x:
  1. Sebelum pengantaran (di SPPG)
  2. Saat tiba di lokasi / sebelum dikonsumsi
• Parameter nilai 1-5:
  - Rasa: 1(tidak enak) - 5(sangat enak)
  - Warna: 1(tidak menarik) - 5(sangat menarik)
  - Aroma: 1(tidak sedap) - 5(sangat wangi)
  - Tekstur: 1(tidak sesuai) - 5(sangat sesuai)
• Form dibawa oleh sopir
• Kesimpulan: Aman / Tidak aman dikonsumsi
• Per satuan pendidikan: 2 paket uji
• Per kelompok 3B: 1 paket uji

*KETENTUAN MAKANAN:*
• Wajib dikonsumsi max 4 jam setelah dimasak
• Jika ada Kejadian Menonjol (KM) gangguan pencernaan
  → laporkan segera
"""

JUKNIS_KEUANGAN = """
💳 *MEKANISME KEUANGAN & DANA SPPG*

*SUMBER DANA:*
• APBN melalui DIPA Badan Gizi Nasional
• Bantuan Pemerintah (Banper)
• Mekanisme: Langsung ke VA (Virtual Account) SPPG

*ALUR PENCARIAN DANA:*
1. Proposal diajukan Yayasan → divalidasi & diverifikasi
2. Dana awal masuk ke VA SPPG: Rp500.000.000
3. Top-up otomatis: seminggu sekali berdasarkan
   laporan harian penggunaan dana
4. Ceiling maksimal saldo VA: Rp500.000.000
5. Penarikan Dana: SPM-LS melalui KPPN

*PEMBAGIAN DANA:*
• Bahan Baku Pangan (sesuai jumlah PM × biaya satuan)
• Biaya Operasional (listrik, gas, air, insentif, dll)
• Insentif Fasilitas SPPG (Rp6jt/hari, dibayar 2 mingguan)

*PETTY CASH (Kas Kecil):*
• Max Rp500.000 per transaksi
• Max Rp5.000.000 per minggu
• Digunakan untuk pengeluaran rutin kecil
• Dicatat di Buku Bantu Petty Cash (Lamp 30f)

*SISA DANA:*
• Sisa dana per periode DIALIHKAN ke periode berikutnya
• Dibuatkan Berita Acara Pengalihan Sisa Dana (Lamp 30n)

*PERTANGGUNGJAWABAN:*
• Laporan 2 mingguan (Lamp 30c)
• Laporan bulanan (Lamp 30d)
• Surat Pernyataan Tanggung Jawab (Lamp 30j)
• Semua pengeluaran harus ada bukti sah (kuitansi)
"""

JUKNIS_SOP = """
⚙️ *SOP OPERASIONAL SPPG*

*ALUR PRODUKSI HARIAN:*
1. Pengadaan/Penerimaan Bahan Baku
   → Sortir & QC oleh Admin Gudang + Aslap
   → Timbang & catat di Buku Bahan Pangan
2. Persiapan Bahan
   → Cuci, potong, sortir ulang
   → QC sayur: periksa ulat, busuk, kualitas
3. Pengolahan (Cooking)
   → Juru masak & tim memasak sesuai menu
   → Cek rasa, kematangan, porsi
4. Pemorsian
   → Timbang & porsikan per kelompok sasaran
   → QC lapisan terakhir sebelum packing
5. Packing & Labeling
   → Masukkan ke ompreng, label sesuai lokasi
6. Distribusi
   → Muat ke kendaraan box alumunium
   → Antar ke satuan Pendidikan/Posyandu
   → Sopir bawa form uji organoleptik
7. Penerimaan di Lokasi
   → Uji organoleptik oleh PIC satuan Pendidikan
   → BAST ditandatangani
   → Makanan dikonsumsi max 4 jam
8. Pencucian & Pengembalian Ompreng
   → Ompreng dikembalikan ke SPPG
   → Dicuci & disterilkan
9. Pencatatan & Pelaporan
   → Data produksi, distribusi, waste, masalah
   → Input ke SIPGN (Sistem Informasi Pemenuhan Gizi Nasional)

*SURVEI HARGA:*
• Dilakukan mingguan oleh Ka.SPPG + Pengawas Keuangan
• Bandingkan minimal 3 penyedia per komoditas
• Acuan HET daerah setempat atau Indeks Kemahalan

*PENANGANAN KEJADIAN MENONJOL (KM):*
1. Pisahkan produk bermasalah
2. Catat detail kejadian
3. Laporkan ke Ka.SPPG & KPPG
4. Investigasi root cause
5. Tindakan korektif
6. Dokumentasi dalam laporan
"""


async def show_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show today's operational status"""
    try:
        today = str(date.today())
        today_file = HARIAN_DIR / f"{today}.json"

        msg = "📈 *STATUS SPPG HARI INI*\n\n"

        if today_file.exists():
            data = load_json(today_file)
            msg += (
                f"📅 {tgl_indo(today)} ({nama_hari(today)})\n"
                f"🏢 {data.get('nama_sppg', 'SPPG')}\n\n"
                f"🎯 Target: {data.get('target', 0)} porsi\n"
                f"🍳 Produksi: {data.get('produksi', 0)} porsi\n"
                f"🚚 Distribusi: {data.get('distribusi', 0)} porsi\n"
                f"✅ Tepat Waktu: {'Ya' if data.get('tepat_waktu') else 'Tidak'}\n"
                f"👥 Total PM: {data.get('total_pm', 0)}\n"
            )
            if data.get('masalah'):
                msg += f"\n⚠️ Masalah: {data['masalah']}\n"
            else:
                msg += "\n✅ Tidak ada masalah\n"
            msg += f"\n📝 {data.get('catatan', '')}"
        else:
            msg += "📅 Belum ada laporan untuk hari ini.\n\n"
            msg += "Ketik /harian untuk mulai input."

        if isinstance(update, Update) and update.callback_query:
            await update.callback_query.edit_message_text(msg, parse_mode="Markdown")
        elif update.message:
            await update.message.reply_text(msg, parse_mode="Markdown")
    except Exception as e:
        await safe_reply(update, f"❌ Error: {e}")


# ═══════════════════════════════════════
#  LAPORAN HARIAN CONVERSATION
# ═══════════════════════════════════════

async def harian_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Start daily report input"""
    context.user_data["harian"] = {}
    today = str(date.today())
    context.user_data["harian"]["tanggal"] = today
    context.user_data["harian"]["nama_sppg"] = "SPPG Wonodri 3"

    await update.message.reply_text(
        f"📋 *LAPORAN HARIAN SPPG*\n"
        f"Tanggal: {tgl_indo(today)} ({nama_hari(today)})\n"
        f"SPPG: SPPG Wonodri 3\n\n"
        f"📦 *PRODUKSI & DISTRIBUSI*\n\n"
        f"Berapa target porsi hari ini?",
        parse_mode="Markdown"
    )
    return STATE_HARIAN_TARGET


async def harian_target(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["harian"]["target"] = int(update.message.text.strip())
        await update.message.reply_text("🍳 Berapa porsi yang diproduksi?")
        return STATE_HARIAN_PRODUKSI
    except:
        await update.message.reply_text("❌ Masukkan angka saja. Contoh: 1408")
        return STATE_HARIAN_TARGET


async def harian_produksi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["harian"]["produksi"] = int(update.message.text.strip())
        await update.message.reply_text("🚚 Berapa porsi yang didistribusikan?")
        return STATE_HARIAN_DISTRIBUSI
    except:
        await update.message.reply_text("❌ Masukkan angka saja.")
        return STATE_HARIAN_PRODUKSI


async def harian_distribusi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["harian"]["distribusi"] = int(update.message.text.strip())
        context.user_data["harian"]["tepat_waktu"] = True
        context.user_data["harian"]["kebersihan"] = "baik"
        await update.message.reply_text(
            "👥 *RINCIAN PENERIMA MANFAAT*\n\n"
            "Jumlah *Besar* (SD 4-6/SMP/SMA/ATS 9-18) — Rp10.000:",
            parse_mode="Markdown"
        )
        return STATE_HARIAN_BESAR
    except:
        await update.message.reply_text("❌ Masukkan angka saja.")
        return STATE_HARIAN_DISTRIBUSI


async def harian_besar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["harian"]["jml_besar"] = int(update.message.text.strip() or 0)
        await update.message.reply_text("Jumlah *Tendik* (Pendidik & Tenaga Kependidikan):", parse_mode="Markdown")
        return STATE_HARIAN_TENDIK
    except:
        await update.message.reply_text("❌ Masukkan angka saja.")
        return STATE_HARIAN_BESAR


async def harian_tendik(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["harian"]["jml_tendik"] = int(update.message.text.strip() or 0)
        await update.message.reply_text("Jumlah *Kecil* (PAUD/TK/SD 1-3/Balita/ATS <9) — Rp8.000:", parse_mode="Markdown")
        return STATE_HARIAN_KECIL
    except:
        await update.message.reply_text("❌ Masukkan angka saja.")
        return STATE_HARIAN_TENDIK


async def harian_kecil(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["harian"]["jml_kecil"] = int(update.message.text.strip() or 0)
        await update.message.reply_text("Jumlah *3B* (Ibu Hamil/Menyusui/Balita 6-59bln):", parse_mode="Markdown")
        return STATE_HARIAN_3B
    except:
        await update.message.reply_text("❌ Masukkan angka saja.")
        return STATE_HARIAN_KECIL


async def harian_3b(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        d = context.user_data["harian"]
        d["jml_3b"] = int(update.message.text.strip() or 0)
        d["total_pm"] = d["jml_besar"] + d["jml_tendik"] + d["jml_kecil"] + d["jml_3b"]
        await update.message.reply_text(
            "🍽️ *MENU HARI INI*\n\n"
            "Karbohidrat (contoh: Nasi Putih):",
            parse_mode="Markdown"
        )
        return STATE_HARIAN_MENU_NASTAR
    except:
        await update.message.reply_text("❌ Masukkan angka saja.")
        return STATE_HARIAN_3B


async def harian_menu_karbohidrat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["harian"]["menu_karbohidrat"] = update.message.text.strip()
    await update.message.reply_text("Protein Hewani (contoh: Ayam/Semur Ayam):")
    return STATE_HARIAN_MENU_PROHE


async def harian_menu_prohe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["harian"]["menu_protein_hewani"] = update.message.text.strip()
    await update.message.reply_text("Protein Nabati (contoh: Tahu/Tempe):")
    return STATE_HARIAN_MENU_PRONA


async def harian_menu_prona(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["harian"]["menu_protein_nabati"] = update.message.text.strip()
    await update.message.reply_text("Sayur (contoh: Sayur Kol Wortel):")
    return STATE_HARIAN_MENU_SAYUR


async def harian_menu_sayur(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["harian"]["menu_sayur"] = update.message.text.strip()
    await update.message.reply_text("Buah (contoh: Jeruk):")
    return STATE_HARIAN_MENU_BUAH


async def harian_menu_buah(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["harian"]["menu_buah"] = update.message.text.strip()
    await update.message.reply_text(
        "🗑️ *WASTE MAKANAN (kg)*\n\n"
        "Nasi / Karbohidrat (kg) — ketik 0 jika tidak ada:",
        parse_mode="Markdown"
    )
    return STATE_HARIAN_WASTE_NASAK


async def harian_waste_nasak(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["harian"]["waste_nasi"] = float(update.message.text.strip() or 0)
        await update.message.reply_text("Sayur (kg):")
        return STATE_HARIAN_WASTE_SAYUR
    except:
        await update.message.reply_text("❌ Masukkan angka (contoh: 18.5)")
        return STATE_HARIAN_WASTE_NASAK


async def harian_waste_sayur(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["harian"]["waste_sayur"] = float(update.message.text.strip() or 0)
        await update.message.reply_text("Lauk / Protein (kg):")
        return STATE_HARIAN_WASTE_LAUK
    except:
        await update.message.reply_text("❌ Masukkan angka.")
        return STATE_HARIAN_WASTE_SAYUR


async def harian_waste_lauk(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["harian"]["waste_lauk"] = float(update.message.text.strip() or 0)
        await update.message.reply_text("Buah (kg):")
        return STATE_HARIAN_WASTE_BUAH
    except:
        await update.message.reply_text("❌ Masukkan angka.")
        return STATE_HARIAN_WASTE_LAUK


async def harian_waste_buah(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        d = context.user_data["harian"]
        d["waste_buah"] = float(update.message.text.strip() or 0)
        d["waste_total"] = round(d["waste_nasi"] + d["waste_sayur"] + d["waste_lauk"] + d["waste_buah"], 2)
        await update.message.reply_text(
            "⚠️ *MASALAH OPERASIONAL*\n\n"
            "Ada masalah? (y/tidak):",
            parse_mode="Markdown"
        )
        return STATE_HARIAN_MASALAH
    except:
        await update.message.reply_text("❌ Masukkan angka.")
        return STATE_HARIAN_WASTE_BUAH


async def harian_masalah(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().lower()
    if text.startswith("y"):
        await update.message.reply_text("Jelaskan masalahnya:")
        context.user_data["harian"]["ada_masalah"] = True
        return STATE_HARIAN_TINDAKAN
    else:
        d = context.user_data["harian"]
        d["masalah"] = ""
        d["tindakan"] = ""
        d["kategori_masalah"] = ""
        d["teratasi"] = "y"
        await update.message.reply_text(
            "💬 *FEEDBACK PENERIMA MANFAAT*\n\n"
            "Ada feedback dari PM? (ketik - jika tidak ada):",
            parse_mode="Markdown"
        )
        return STATE_HARIAN_FEEDBACK


async def harian_tindakan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["harian"]["masalah"] = update.message.text.strip()
    await update.message.reply_text("Tindakan yang diambil:")
    # Special state: we need masalah text and then tindakan
    # But we already used masalah state... let's use context
    context.user_data["harian"]["_next_after_tindakan"] = True
    return STATE_HARIAN_TINDAKAN


# Since we have a flow issue with masalah+tindakan, let me handle this differently
# We'll use the same state for kedua input

async def harian_catatan_masalah(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """After explaining masalah, get tindakan, then feedback"""
    # If this is the first time in this state, we got the masalah desc
    # We'll use a flag
    d = context.user_data["harian"]
    if "_got_tindakan" not in d:
        # This is the masalah description
        d["masalah"] = update.message.text.strip()
        d["_got_tindakan"] = True
        await update.message.reply_text("Tindakan yang diambil:")
        return STATE_HARIAN_TINDAKAN
    else:
        # This is the tindakan text
        d["tindakan"] = update.message.text.strip()
        d["kategori_masalah"] = "operasional"
        d["teratasi"] = "y"
        del d["_got_tindakan"]
        await update.message.reply_text(
            "💬 *FEEDBACK PENERIMA MANFAAT*\n\n"
            "Ada feedback dari PM? (ketik - jika tidak ada):",
            parse_mode="Markdown"
        )
        return STATE_HARIAN_FEEDBACK


async def harian_feedback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    context.user_data["harian"]["feedback_pm"] = "" if text == "-" else text
    await update.message.reply_text(
        "📝 *CATATAN TAMBAHAN*\n\n"
        "Ada catatan? (ketik - jika tidak ada):",
        parse_mode="Markdown"
    )
    return STATE_HARIAN_CATATAN


async def harian_catatan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    context.user_data["harian"]["catatan"] = "" if text == "-" else text

    # Save and generate report
    try:
        d = context.user_data["harian"]
        tgl = d["tanggal"]

        # Save JSON
        json_path = HARIAN_DIR / f"{tgl}.json"
        save_json(json_path, d)

        # Generate Excel
        xlsx_path = HARIAN_DIR / f"laporan_harian_{tgl}.xlsx"
        _make_harian_excel(d, xlsx_path)

        # Send to user
        total_waste = d.get("waste_total", 0)

        msg = (
            f"✅ *LAPORAN HARIAN TERSIMPAN*\n\n"
            f"📅 {tgl_indo(tgl)} ({nama_hari(tgl)})\n"
            f"🏢 {d.get('nama_sppg', 'SPPG')}\n\n"
            f"🎯 Target: {d['target']} → Produksi: {d['produksi']} → Distribusi: {d['distribusi']}\n"
            f"👥 Total PM: {d['total_pm']}\n"
            f"🗑️ Waste: {total_waste} kg\n"
        )
        if d.get("masalah"):
            msg += f"⚠️ Masalah: {d['masalah'][:50]}...\n"
        else:
            msg += "✅ Tidak ada masalah\n"

        await update.message.reply_text(msg, parse_mode="Markdown")

        # Send Excel
        if xlsx_path.exists():
            doc = FSInputFile(xlsx_path)
            await update.message.reply_document(
                document=doc,
                filename=f"laporan_harian_{tgl}.xlsx",
                caption=f"📋 Laporan Harian {tgl_indo(tgl)}"
            )

    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")

    # Back to main menu
    keyboard = [[InlineKeyboardButton("🔙 Kembali ke Menu", callback_data="back_menu")]]
    await update.message.reply_text(
        "Pilih menu selanjutnya:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

    return ConversationHandler.END


def _make_harian_excel(data, filepath):
    """Generate simple daily report Excel"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Harian"

    tgl = data["tanggal"]
    ws.merge_cells("A1:F1")
    ws["A1"].value = f"LAPORAN OPERASIONAL HARIAN — {tgl_indo(tgl)}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:F2")
    ws["A2"].value = f"{data.get('nama_sppg', 'SPPG')} | {nama_hari(tgl)}"
    ws["A2"].font = Font(name="Calibri", size=11, color="555555")
    ws["A2"].alignment = CENTER

    ROW = 4
    sections = [
        ("📦 PRODUKSI", [
            ("Target", f"{data['target']} porsi"),
            ("Produksi", f"{data['produksi']} porsi"),
            ("Distribusi", f"{data['distribusi']} porsi"),
            ("% Pencapaian", f"{(data['produksi']/data['target']*100) if data['target'] else 0:.0f}%"),
        ]),
        ("👥 PENERIMA MANFAAT", [
            ("Besar", str(data.get("jml_besar", 0))),
            ("Tendik", str(data.get("jml_tendik", 0))),
            ("Kecil", str(data.get("jml_kecil", 0))),
            ("3B", str(data.get("jml_3b", 0))),
            ("TOTAL", str(data.get("total_pm", 0))),
        ]),
        ("🍽️ MENU", [
            ("Karbohidrat", data.get("menu_karbohidrat", "")),
            ("Protein Hewani", data.get("menu_protein_hewani", "")),
            ("Protein Nabati", data.get("menu_protein_nabati", "")),
            ("Sayur", data.get("menu_sayur", "")),
            ("Buah", data.get("menu_buah", "")),
        ]),
        ("🗑️ WASTE (kg)", [
            ("Nasi", f"{data.get('waste_nasi', 0)} kg"),
            ("Sayur", f"{data.get('waste_sayur', 0)} kg"),
            ("Lauk", f"{data.get('waste_lauk', 0)} kg"),
            ("Buah", f"{data.get('waste_buah', 0)} kg"),
            ("TOTAL", f"{data.get('waste_total', 0)} kg"),
        ]),
    ]

    for title, items in sections:
        ws.merge_cells(f"A{ROW}:F{ROW}")
        ws[f"A{ROW}"].value = title
        ws[f"A{ROW}"].font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
        ws[f"A{ROW}"].fill = SUBHEADER
        ROW += 1
        for label, val in items:
            ws.cell(ROW, 1, label).font = Font(name="Calibri", bold=True, size=10)
            ws.cell(ROW, 2, val).font = Font(name="Calibri", size=10)
            ws.cell(ROW, 1).fill = LIGHT_GRAY
            for c in range(1, 4):
                ws.cell(ROW, c).border = THIN
            ROW += 1
        ROW += 1

    # Masalah
    ws.merge_cells(f"A{ROW}:F{ROW}")
    ws[f"A{ROW}"].value = "⚠️  MASALAH"
    ws[f"A{ROW}"].font = Font(name="Calibri", bold=True, size=11, color="C00000")
    ws[f"A{ROW}"].fill = PatternFill("solid", fgColor="FFF2CC")
    ROW += 1
    if data.get("masalah"):
        ws.cell(ROW, 1, data["masalah"]).font = FONT_BODY
        ws.merge_cells(f"A{ROW}:F{ROW}")
    else:
        ws[f"A{ROW}"].value = "✅ Tidak ada masalah"
        ws[f"A{ROW}"].font = FONT_BODY
        ws[f"A{ROW}"].fill = GREEN_FILL

    ROW += 3
    ws[f"A{ROW}"].value = f"Disusun oleh: Bot SPPG | {tgl_indo(tgl)}"
    ws[f"A{ROW}"].font = Font(name="Calibri", italic=True, size=9, color="999999")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    wb.save(filepath)


# ═══════════════════════════════════════
#  LAPORAN MINGGUAN CONVERSATION
# ═══════════════════════════════════════

async def mingguan_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Start weekly report input manual"""
    context.user_data["mingguan"] = {}
    context.user_data["mingguan"]["nama_sppg"] = "SPPG Wonodri 3"
    context.user_data["mingguan"]["penyusun"] = "Asisten Lapangan"

    await update.message.reply_text(
        "📊 *LAPORAN MINGGUAN*\n\n"
        "Input periode (contoh: 2026-06-15):\n"
        "Tanggal *mulai*:",
        parse_mode="Markdown"
    )
    return STATE_MINGGUAN_TGL_START


async def mingguan_tgl_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["mingguan"]["tgl_start"] = update.message.text.strip()
    await update.message.reply_text("Tanggal *selesai* (contoh: 2026-06-19):", parse_mode="Markdown")
    return STATE_MINGGUAN_TGL_END


async def mingguan_tgl_end(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["mingguan"]["tgl_end"] = update.message.text.strip()

    # Load daily data for this range
    from auto_summary import load_week_data as load_data
    days, start_d, end_d = load_data(
        context.user_data["mingguan"]["tgl_start"],
        context.user_data["mingguan"]["tgl_end"]
    )
    aktif_days = [d for d in days if d.get("produksi", 0) > 0]

    if not aktif_days:
        await update.message.reply_text("❌ Tidak ada data harian di periode ini. Isi laporan harian dulu.")
        return ConversationHandler.END

    context.user_data["mingguan"]["days"] = days
    context.user_data["mingguan"]["tgl_start"] = str(start_d)
    context.user_data["mingguan"]["tgl_end"] = str(end_d)
    context.user_data["mingguan"]["tgl_laporan"] = str(date.today())

    total_prod = sum(d["produksi"] for d in aktif_days)
    await update.message.reply_text(
        f"✅ Ditemukan {len(aktif_days)} hari data aktif\n"
        f"Total produksi: {total_prod:,} porsi\n\n"
        f"📈 *ROOT CAUSE ANALYSIS*\n\n"
        f"Temuan utama minggu ini:",
        parse_mode="Markdown"
    )
    return STATE_MINGGUAN_ANALISA_TEMUAN


async def mingguan_analisa_temuan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["mingguan"]["analisa_temuan"] = update.message.text.strip()
    await update.message.reply_text("Root cause:")
    return STATE_MINGGUAN_ANALISA_ROOT


async def mingguan_analisa_root(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["mingguan"]["analisa_root"] = update.message.text.strip()
    await update.message.reply_text("Tindakan korektif:")
    return STATE_MINGGUAN_ANALISA_TINDAKAN


async def mingguan_analisa_tindakan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["mingguan"]["analisa_tindakan"] = update.message.text.strip()
    await update.message.reply_text(
        "🎯 *REKOMENDASI*\n\n"
        "Masukkan rekomendasi (1 per baris).\n"
        "Ketik SELESAI jika sudah cukup:",
        parse_mode="Markdown"
    )
    context.user_data["mingguan"]["rekomendasi"] = []
    return STATE_MINGGUAN_REKOMENDASI


async def mingguan_rekomendasi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if text.upper() == "SELESAI":
        await update.message.reply_text(
            "🔥 *EXECUTIVE SUMMARY*\n\n"
            "Tulis ringkasan eksekutif minggu ini:",
            parse_mode="Markdown"
        )
        return STATE_MINGGUAN_EXEC_SUMMARY
    else:
        context.user_data["mingguan"]["rekomendasi"].append(text)
        await update.message.reply_text(f"✅ Rekomendasi #{len(context.user_data['mingguan']['rekomendasi'])} disimpan. Kirim lagi atau ketik SELESAI.")
        return STATE_MINGGUAN_REKOMENDASI


async def mingguan_exec_summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["mingguan"]["exec_summary"] = update.message.text.strip()
    await update.message.reply_text(
        "💡 *INSIGHT & INPUTAN*\n\n"
        "Ada insight tambahan? (ketik - jika tidak):",
        parse_mode="Markdown"
    )
    return STATE_MINGGUAN_INSIGHT


async def mingguan_insight(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    d = context.user_data["mingguan"]

    # Build report
    from laporan_mingguan import render_excel
    from auto_summary import auto_evaluasi_divisi, auto_summary

    report = {
        "nama_sppg": d.get("nama_sppg", "SPPG Wonodri 3"),
        "penyusun": d.get("penyusun", "Asisten Lapangan"),
        "tgl_start": d["tgl_start"],
        "tgl_end": d["tgl_end"],
        "tgl_laporan": str(date.today()),
        "days": d["days"],
        "analisa": {
            "temuan": d.get("analisa_temuan", ""),
            "root_cause": d.get("analisa_root", ""),
            "tindakan_korektif": d.get("analisa_tindakan", ""),
        },
        "evaluasi_divisi": auto_evaluasi_divisi(d["days"]),
        "rekomendasi": d.get("rekomendasi", []),
        "executive_summary": d.get("exec_summary", ""),
        "insight": "" if text == "-" else text,
        "status_performa": "Baik",
    }

    # Save
    tgl_key = f"{report['tgl_start']}_to_{report['tgl_end']}"
    json_path = MINGGUAN_DIR / f"mingguan_{tgl_key}.json"
    save_json(json_path, report)

    xlsx_path = MINGGUAN_DIR / f"laporan_mingguan_{tgl_key}.xlsx"
    render_excel(report, xlsx_path)

    total_prod = sum(d["produksi"] for d in report["days"] if d.get("produksi", 0) > 0)

    msg = (
        f"✅ *LAPORAN MINGGUAN TERSIMPAN*\n\n"
        f"📅 {tgl_indo(report['tgl_start'])} – {tgl_indo(report['tgl_end'])}\n"
        f"🏢 {report['nama_sppg']}\n"
        f"📊 Total Produksi: {total_prod:,} porsi\n"
        f"📝 Rekomendasi: {len(report['rekomendasi'])} item\n"
    )

    await update.message.reply_text(msg, parse_mode="Markdown")

    if xlsx_path.exists():
        doc = FSInputFile(xlsx_path)
        await update.message.reply_document(
            document=doc,
            filename=f"laporan_mingguan_{tgl_key}.xlsx",
            caption=f"📊 Weekly Report {tgl_indo(report['tgl_start'])} - {tgl_indo(report['tgl_end'])}"
        )

    keyboard = [[InlineKeyboardButton("🔙 Kembali ke Menu", callback_data="back_menu")]]
    await update.message.reply_text("Selesai!", reply_markup=InlineKeyboardMarkup(keyboard))

    return ConversationHandler.END


# ═══════════════════════════════════════
#  AUTO MINGGUAN (from daily data)
# ═══════════════════════════════════════

async def mingguan_auto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Auto-generate weekly report from daily data"""
    try:
        end_date = date.today()
        # Find Monday of current week
        start_date = end_date - timedelta(days=end_date.weekday())

        from auto_summary import main as auto_main

        # Load data
        from auto_summary import (
            load_week_data, auto_analisa, auto_evaluasi_divisi,
            auto_rekomendasi, auto_summary
        )
        from laporan_mingguan import render_excel

        days, start_d, end_d = load_week_data(str(start_date), str(end_date))
        aktif_days = [d for d in days if d.get("produksi", 0) > 0]

        if not aktif_days:
            await update.message.reply_text(
                "❌ Tidak ada data harian di minggu ini.\n"
                "Isi laporan harian dulu via /harian"
            )
            return

        # Auto-generate
        analisa = auto_analisa(aktif_days)
        evaluasi = auto_evaluasi_divisi(aktif_days)
        rekom = auto_rekomendasi(aktif_days, analisa)
        summary = auto_summary(aktif_days, analisa)

        report = {
            "nama_sppg": "SPPG Wonodri 3",
            "penyusun": "Auto-Generated Bot",
            "tgl_start": str(start_d),
            "tgl_end": str(end_d),
            "tgl_laporan": str(date.today()),
            "days": days,
            "analisa": analisa,
            "evaluasi_divisi": evaluasi,
            "rekomendasi": rekom,
            "executive_summary": summary,
            "insight": "Auto-generated from daily report data via Telegram bot.",
            "status_performa": "Baik",
        }

        tgl_key = f"{report['tgl_start']}_to_{report['tgl_end']}"
        json_path = MINGGUAN_DIR / f"auto_mingguan_{tgl_key}.json"
        save_json(json_path, report)

        xlsx_path = MINGGUAN_DIR / f"auto_laporan_mingguan_{tgl_key}.xlsx"
        render_excel(report, xlsx_path)

        total_prod = sum(d["produksi"] for d in aktif_days)
        waste_total = sum(d.get("waste_total", 0) for d in aktif_days)

        msg = (
            f"✅ *AUTO-SUMMARY MINGGUAN*\n\n"
            f"📅 {tgl_indo(str(start_d))} – {tgl_indo(str(end_d))}\n"
            f"📊 Total Produksi: {total_prod:,} porsi\n"
            f"🗑️ Total Waste: {waste_total:.1f} kg\n"
            f"⚠️ Temuan: {analisa.get('temuan', '-')[:80]}\n"
            f"🎯 Rekomendasi: {len(rekom)} item\n"
            f"📈 Status: {report['status_performa']}\n"
        )

        await update.message.reply_text(msg, parse_mode="Markdown")

        if xlsx_path.exists():
            doc = FSInputFile(xlsx_path)
            await update.message.reply_document(
                document=doc,
                filename=f"auto_mingguan_{tgl_key}.xlsx",
                caption=f"📊 Auto Weekly Report"
            )

    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


# ═══════════════════════════════════════
#  PENERIMA MANFAAT
# ═══════════════════════════════════════

async def penerima_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Start penerima manfaat report"""
    context.user_data["penerima"] = {}

    keyboard = [
        [InlineKeyboardButton("Google Sheets URL", callback_data="pm_gsheet")],
        [InlineKeyboardButton("Upload CSV", callback_data="pm_csv")],
        [InlineKeyboardButton("Input Manual", callback_data="pm_manual")],
    ]
    await update.message.reply_text(
        "📑 *REPORT PENERIMA MANFAAT*\n\n"
        "Pilih sumber data:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )
    return STATE_PENERIMA_SOURCE


async def penerima_source_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "pm_gsheet":
        await query.edit_message_text(
            "📎 Kirim link Google Sheets-nya.\n\n"
            "Format: https://docs.google.com/spreadsheets/d/...\n\n"
            "Csv export juga bisa langsung."
        )
        return STATE_PENERIMA_URL
    elif query.data == "pm_csv":
        await query.edit_message_text(
            "📤 Upload file CSV.\n\n"
            "Format kolom: Waktu,Tempat,Besar,Tendik,Kecil"
        )
        return STATE_PENERIMA_ROW
    elif query.data == "pm_manual":
        await query.edit_message_text(
            "✏️ Input manual.\n\n"
            "Format per baris:\n"
            "<waktu>,<tempat>,<besar>,<tendik>,<kecil>\n\n"
            "Contoh:\n"
            "07.00,SDN Lamper Lor,79,12,63\n"
            "07.00,TK Sula,0,15,0\n\n"
            "Ketik SELESAI jika sudah cukup."
        )
        return STATE_PENERIMA_ROW


async def penerima_url(update: Update, context: ContextTypes.DEFAULT_TYPE):
    url = update.message.text.strip()
    from report_penerima import fetch_from_gsheet

    rows = fetch_from_gsheet(url)
    if not rows:
        await update.message.reply_text("❌ Gagal fetch data. Cek URL-nya.")
        return STATE_PENERIMA_URL

    # Process
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = PENERIMA_DIR / f"penerima_{timestamp}.json"
    save_json(json_path, rows)

    from report_penerima import render_excel
    xlsx_path = PENERIMA_DIR / f"report_penerima_{timestamp}.xlsx"
    render_excel(rows, "Google Sheets", xlsx_path, "SPPG Wonodri 3")

    total_pm = sum(int(r.get("Total", 0)) for r in rows if r.get("Total", ""))
    await update.message.reply_text(
        f"✅ *REPORT PENERIMA MANFAAT*\n\n"
        f"📊 Total PM: {total_pm:,} orang\n"
        f"🏫 Institusi: {len(rows)} tempat",
        parse_mode="Markdown"
    )

    if xlsx_path.exists():
        doc = FSInputFile(xlsx_path)
        await update.message.reply_document(
            document=doc,
            filename=f"report_penerima_{timestamp}.xlsx",
            caption="📋 Data Penerima Manfaat SPPG"
        )

    keyboard = [[InlineKeyboardButton("🔙 Kembali ke Menu", callback_data="back_menu")]]
    await update.message.reply_text("Selesai!", reply_markup=InlineKeyboardMarkup(keyboard))

    return ConversationHandler.END


async def penerima_rows(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle manual CSV input"""
    text = update.message.text.strip()

    # Check for file upload
    if update.message.document:
        file = update.message.document
        if file.file_name.endswith(".csv"):
            import tempfile
            import csv

            tf = await file.get_file()
            content = await tf.download_as_bytearray()
            content_str = content.decode("utf-8")
            reader = csv.DictReader(io.StringIO(content_str))
            rows = list(reader)

            if rows:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                json_path = PENERIMA_DIR / f"penerima_{timestamp}.json"
                save_json(json_path, rows)

                from report_penerima import render_excel
                xlsx_path = PENERIMA_DIR / f"report_penerima_{timestamp}.xlsx"
                render_excel(rows, "CSV Upload", xlsx_path, "SPPG Wonodri 3")

                total_pm = sum(int(r.get("Total", 0)) for r in rows if r.get("Total", ""))
                await update.message.reply_text(
                    f"✅ *REPORT PENERIMA MANFAAT*\n\n"
                    f"📊 Total PM: {total_pm:,} orang\n"
                    f"🏫 Institusi: {len(rows)} tempat",
                    parse_mode="Markdown"
                )

                if xlsx_path.exists():
                    doc = FSInputFile(xlsx_path)
                    await update.message.reply_document(document=doc, filename=f"report_penerima_{timestamp}.xlsx")

                keyboard = [[InlineKeyboardButton("🔙 Kembali ke Menu", callback_data="back_menu")]]
                await update.message.reply_text("Selesai!", reply_markup=InlineKeyboardMarkup(keyboard))

                return ConversationHandler.END

        await update.message.reply_text("❌ File harus CSV. Format: Waktu,Tempat,Besar,Tendik,Kecil")
        return STATE_PENERIMA_ROW

    # Manual input
    if text.upper() == "SELESAI":
        rows = context.user_data.get("penerima_rows", [])
        if not rows:
            await update.message.reply_text("❌ Tidak ada data.")
            return ConversationHandler.END

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_path = PENERIMA_DIR / f"penerima_{timestamp}.json"
        save_json(json_path, rows)

        from report_penerima import render_excel
        xlsx_path = PENERIMA_DIR / f"report_penerima_{timestamp}.xlsx"
        render_excel(rows, "Input Manual", xlsx_path, "SPPG Wonodri 3")

        total_pm = sum(int(r.get("Total", 0)) for r in rows if r.get("Total", ""))
        await update.message.reply_text(
            f"✅ *REPORT PENERIMA MANFAAT*\n\n"
            f"📊 Total PM: {total_pm:,} orang\n"
            f"🏫 Institusi: {len(rows)} tempat",
            parse_mode="Markdown"
        )

        if xlsx_path.exists():
            doc = FSInputFile(xlsx_path)
            await update.message.reply_document(document=doc, filename=f"report_penerima_{timestamp}.xlsx")

        keyboard = [[InlineKeyboardButton("🔙 Kembali ke Menu", callback_data="back_menu")]]
        await update.message.reply_text("Selesai!", reply_markup=InlineKeyboardMarkup(keyboard))

        return ConversationHandler.END
    else:
        # Parse CSV line
        parts = [p.strip() for p in text.split(",")]
        if len(parts) >= 5:
            row = {
                "Waktu": parts[0],
                "Tempat": parts[1],
                "Besar": parts[2],
                "Tendik": parts[3],
                "Kecil": parts[4],
                "Total": str(int(parts[2]) + int(parts[3]) + int(parts[4]))
            }
            if "penerima_rows" not in context.user_data:
                context.user_data["penerima_rows"] = []
            context.user_data["penerima_rows"].append(row)
            await update.message.reply_text(
                f"✅ {parts[1]}: Total {row['Total']} PM. Kirim lagi atau ketik SELESAI."
            )
        else:
            await update.message.reply_text(
                "❌ Format salah. Gunakan:\n"
                "<waktu>,<tempat>,<besar>,<tendik>,<kecil>\n"
                "Contoh: 07.00,SDN Lamper Lor,79,12,63"
            )
        return STATE_PENERIMA_ROW


# ═══════════════════════════════════════
#  FALLBACK & CANCEL
# ═══════════════════════════════════════

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ Dibatalkan.")
    return ConversationHandler.END



async def menu_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle menu button callbacks — main navigation"""
    query = update.callback_query
    await query.answer()
    action = query.data

    if action == "menu_harian":
        await query.edit_message_text(
            "📋 *LAPORAN HARIAN*\n\n"
            "Ketik /harian untuk memulai input laporan operasional harian.\n\n"
            "Data yang diisi:\n"
            "• Target & Produksi\n"
            "• Rincian penerima manfaat\n"
            "• Menu hari ini\n"
            "• Waste makanan (kg)\n"
            "• Masalah & feedback PM\n"
            "• Catatan tambahan",
            parse_mode="Markdown"
        )
    elif action == "menu_mingguan":
        await query.edit_message_text(
            "📊 *LAPORAN MINGGUAN*\n\n"
            "Ada 2 cara:\n\n"
            "1️⃣ *Auto-Summary* (langsung dari data harian)\n"
            "   Ketik: /mingguan_auto\n\n"
            "2️⃣ *Manual* (analisa + rekomendasi sendiri)\n"
            "   Ketik: /mingguan\n\n"
            "Pilih sesuai kebutuhan.",
            parse_mode="Markdown"
        )
    elif action == "menu_penerima":
        await query.edit_message_text(
            "📑 *REPORT PENERIMA MANFAAT*\n\n"
            "Ketik /penerima untuk update data penerima manfaat.\n\n"
            "Bisa dari:\n"
            "• Link Google Sheets\n"
            "• Upload file CSV\n"
            "• Input manual\n\n"
            "Format kolom: Waktu, Tempat, Besar, Tendik, Kecil",
            parse_mode="Markdown"
        )
    elif action == "menu_juknis":
        keyboard = [
            [InlineKeyboardButton("🏢 Struktur Organisasi", callback_data="juknis_struktur")],
            [InlineKeyboardButton("💰 Biaya Acuan", callback_data="juknis_biaya")],
            [InlineKeyboardButton("📅 Hari Operasional", callback_data="juknis_hari")],
            [InlineKeyboardButton("👥 Penerima Manfaat", callback_data="juknis_penerima")],
            [InlineKeyboardButton("📋 Laporan Wajib", callback_data="juknis_laporan")],
            [InlineKeyboardButton("🔬 Standar Gizi & Organoleptik", callback_data="juknis_gizi")],
            [InlineKeyboardButton("💳 Keuangan & Dana", callback_data="juknis_keuangan")],
            [InlineKeyboardButton("⚙️ SOP Operasional", callback_data="juknis_sop")],
            [InlineKeyboardButton("📦 Produksi & Distribusi", callback_data="juknis_produksi")],
            [InlineKeyboardButton("📊 Monitoring & Evaluasi", callback_data="juknis_monitoring")],
            [InlineKeyboardButton("🔙 Kembali", callback_data="back_menu")],
        ]
        await query.edit_message_text(
            "📚 *JUKNIS SPPG MBG 2026*\n"
            "Keputusan Kepala BGN No. 401.1 Tahun 2025\n\n"
            "Pilih topik:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode="Markdown"
        )
    elif action == "menu_status":
        await show_status(update, context)
    elif action.startswith("juknis_"):
        content_map = {
            "juknis_struktur": JUKNIS_STRUKTUR,
            "juknis_biaya": JUKNIS_BIAYA,
            "juknis_hari": JUKNIS_HARI,
            "juknis_penerima": JUKNIS_PENERIMA,
            "juknis_laporan": JUKNIS_LAPORAN,
            "juknis_gizi": JUKNIS_GIZI,
            "juknis_keuangan": JUKNIS_KEUANGAN,
            "juknis_sop": JUKNIS_SOP,
            "juknis_produksi": (
                "📦 *PRODUKSI & DISTRIBUSI*\n"
                "\n"
                "*ALUR PRODUKSI HARIAN:*\n"
                "1. Penerimaan Bahan Baku\n"
                "   → Sortir & QC oleh Admin Gudang + Asisten Lapangan\n"
                "   → Timbang & catat di Buku Bahan Pangan\n"
                "2. Persiapan Bahan\n"
                "   → Cuci, potong, sortir ulang\n"
                "   → QC sayur: periksa ulat, busuk, kualitas\n"
                "3. Pengolahan (Cooking)\n"
                "   → Juru masak & tim memasak sesuai menu\n"
                "   → Cek rasa, kematangan, porsi\n"
                "4. Pemorsian\n"
                "   → Timbang & porsikan per kelompok sasaran\n"
                "   → QC lapisan terakhir sebelum packing\n"
                "5. Packing & Labeling\n"
                "   → Masukkan ke ompreng, label sesuai lokasi\n"
                "6. Distribusi\n"
                "   → Muat ke kendaraan box alumunium\n"
                "   → Antar ke satuan Pendidikan/Posyandu\n"
                "   → Sopir bawa form uji organoleptik\n"
                "7. Penerimaan di Lokasi\n"
                "   → Uji organoleptik oleh PIC satuan Pendidikan\n"
                "   → BAST ditandatangani\n"
                "   → Makanan dikonsumsi max 4 jam\n"
                "8. Pencucian & Pengembalian Ompreng\n"
                "   → Ompreng dikembalikan ke SPPG\n"
                "   → Dicuci & disterilkan\n"
            ),
            "juknis_monitoring": (
                "📊 *MONITORING & EVALUASI*\n"
                "\n"
                "*MONITORING HARIAN:*\n"
                "• Data produksi, distribusi, waste, masalah\n"
                "• Input ke SIPGN setiap hari\n"
                "• Monitoring petty cash & realisasi insentif relawan\n"
                "\n"
                "*EVALUASI BERKALA:*\n"
                "• Evaluasi menu: berdasarkan sisa/waste & feedback penerima\n"
                "• Evaluasi kinerja relawan per shift\n"
                "• Evaluasi keuangan: realisasi vs anggaran bulanan\n"
                "\n"
                "*KEJADIAN MENONJOL (KM):*\n"
                "1. Pisahkan produk bermasalah\n"
                "2. Catat detail kejadian\n"
                "3. Laporkan ke Ka.SPPG & KPPG\n"
                "4. Investigasi root cause\n"
                "5. Tindakan korektif\n"
                "6. Dokumentasi dalam laporan bulanan\n"
            ),
        }
        content = content_map.get(action, "")
        if content:
            keyboard = [[InlineKeyboardButton("🔙 Kembali ke Menu Juknis", callback_data="menu_juknis")]]
            await query.edit_message_text(
                content,
                reply_markup=InlineKeyboardMarkup(keyboard),
                parse_mode="Markdown"
            )


async def back_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    keyboard = [
        [InlineKeyboardButton("📋 Laporan Harian", callback_data="menu_harian")],
        [InlineKeyboardButton("📊 Laporan Mingguan", callback_data="menu_mingguan")],
        [InlineKeyboardButton("📑 Report Penerima Manfaat", callback_data="menu_penerima")],
        [InlineKeyboardButton("📚 Juknis Quick Ref", callback_data="menu_juknis")],
        [InlineKeyboardButton("📈 Status Hari Ini", callback_data="menu_status")],
    ]
    await query.edit_message_text(
        "🍽️ *SPPG REPORT TOOLS*\n\n"
        "Pilih menu:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🤖 *SPPG Bot — Bantuan*\n\n"
        "/start — Menu utama\n"
        "/harian — Input laporan harian\n"
        "/mingguan — Input laporan mingguan (manual)\n"
        "/mingguan_auto — Generate mingguan otomatis\n"
        "/penerima — Report penerima manfaat\n"
        "/juknis — Akses ringkasan JUKNIS\n"
        "/status — Status hari ini\n"
        "/help — Bantuan ini\n"
        "/cancel — Batalkan sesi saat ini",
        parse_mode="Markdown"
    )


async def juknis(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("🏢 Struktur Organisasi", callback_data="juknis_struktur")],
        [InlineKeyboardButton("💰 Biaya Acuan", callback_data="juknis_biaya")],
        [InlineKeyboardButton("📅 Hari Operasional", callback_data="juknis_hari")],
        [InlineKeyboardButton("👥 Penerima Manfaat", callback_data="juknis_penerima")],
        [InlineKeyboardButton("📋 Laporan Wajib", callback_data="juknis_laporan")],
        [InlineKeyboardButton("🔬 Standar Gizi & Organoleptik", callback_data="juknis_gizi")],
        [InlineKeyboardButton("💳 Keuangan & Dana", callback_data="juknis_keuangan")],
        [InlineKeyboardButton("⚙️ SOP Operasional", callback_data="juknis_sop")],
        [InlineKeyboardButton("📦 Produksi & Distribusi", callback_data="juknis_produksi")],
        [InlineKeyboardButton("📊 Monitoring & Evaluasi", callback_data="juknis_monitoring")],
        [InlineKeyboardButton("🔙 Menu Utama", callback_data="back_menu")],
    ]
    await update.message.reply_text(
        JUKNIS_MENU,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )


async def safe_reply(update, text, **kwargs):
    """Reply to either message or callback query"""
    if update.callback_query:
        await update.callback_query.edit_message_text(text, **kwargs)
    elif update.message:
        await update.message.reply_text(text, **kwargs)


# ═══════════════════════════════════════
#  ERROR HANDLER
# ═══════════════════════════════════════

async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle errors"""
    try:
        raise context.error
    except Exception as e:
        error_msg = f"❌ Error: {type(e).__name__}: {str(e)[:200]}"
        try:
            if update and update.effective_chat:
                await context.bot.send_message(
                    chat_id=update.effective_chat.id,
                    text=error_msg
                )
        except:
            pass


# ═══════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════

def main():
    """Start the bot"""
    token = os.environ.get("SPPG_BOT_TOKEN")
    if not token:
        print("❌ SPPG_BOT_TOKEN environment variable not set!")
        sys.exit(1)

    # Create application
    app = (
        Application.builder()
        .token(token)
        .connect_timeout(30.0)
        .read_timeout(30.0)
        .get_updates_connect_timeout(30.0)
        .get_updates_read_timeout(30.0)
        .pool_timeout(30.0)
        .build()
    )

    # ── Simple commands ──
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("juknis", juknis))
    app.add_handler(CommandHandler("status", show_status))
    app.add_handler(CommandHandler("mingguan_auto", mingguan_auto))
    app.add_handler(CallbackQueryHandler(back_menu, pattern="^back_menu$"))
    app.add_handler(CallbackQueryHandler(menu_callback, pattern="^(menu_|juknis_)"))

    # ── Harian conversation ──
    harian_conv = ConversationHandler(
        entry_points=[CommandHandler("harian", harian_start)],
        states={
            STATE_HARIAN_TARGET: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_target)],
            STATE_HARIAN_PRODUKSI: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_produksi)],
            STATE_HARIAN_DISTRIBUSI: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_distribusi)],
            STATE_HARIAN_BESAR: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_besar)],
            STATE_HARIAN_TENDIK: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_tendik)],
            STATE_HARIAN_KECIL: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_kecil)],
            STATE_HARIAN_3B: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_3b)],
            STATE_HARIAN_MENU_NASTAR: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_menu_karbohidrat)],
            STATE_HARIAN_MENU_PROHE: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_menu_prohe)],
            STATE_HARIAN_MENU_PRONA: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_menu_prona)],
            STATE_HARIAN_MENU_SAYUR: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_menu_sayur)],
            STATE_HARIAN_MENU_BUAH: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_menu_buah)],
            STATE_HARIAN_WASTE_NASAK: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_waste_nasak)],
            STATE_HARIAN_WASTE_SAYUR: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_waste_sayur)],
            STATE_HARIAN_WASTE_LAUK: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_waste_lauk)],
            STATE_HARIAN_WASTE_BUAH: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_waste_buah)],
            STATE_HARIAN_MASALAH: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_masalah)],
            STATE_HARIAN_TINDAKAN: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_catatan_masalah)],
            STATE_HARIAN_FEEDBACK: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_feedback)],
            STATE_HARIAN_CATATAN: [MessageHandler(filters.TEXT & ~filters.COMMAND, harian_catatan)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
    )
    app.add_handler(harian_conv)

    # ── Mingguan conversation ──
    mingguan_conv = ConversationHandler(
        entry_points=[CommandHandler("mingguan", mingguan_start)],
        states={
            STATE_MINGGUAN_TGL_START: [MessageHandler(filters.TEXT & ~filters.COMMAND, mingguan_tgl_start)],
            STATE_MINGGUAN_TGL_END: [MessageHandler(filters.TEXT & ~filters.COMMAND, mingguan_tgl_end)],
            STATE_MINGGUAN_ANALISA_TEMUAN: [MessageHandler(filters.TEXT & ~filters.COMMAND, mingguan_analisa_temuan)],
            STATE_MINGGUAN_ANALISA_ROOT: [MessageHandler(filters.TEXT & ~filters.COMMAND, mingguan_analisa_root)],
            STATE_MINGGUAN_ANALISA_TINDAKAN: [MessageHandler(filters.TEXT & ~filters.COMMAND, mingguan_analisa_tindakan)],
            STATE_MINGGUAN_REKOMENDASI: [MessageHandler(filters.TEXT & ~filters.COMMAND, mingguan_rekomendasi)],
            STATE_MINGGUAN_EXEC_SUMMARY: [MessageHandler(filters.TEXT & ~filters.COMMAND, mingguan_exec_summary)],
            STATE_MINGGUAN_INSIGHT: [MessageHandler(filters.TEXT & ~filters.COMMAND, mingguan_insight)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
    )
    app.add_handler(mingguan_conv)

    # ── Penerima conversation ──
    penerima_conv = ConversationHandler(
        entry_points=[CommandHandler("penerima", penerima_start)],
        states={
            STATE_PENERIMA_SOURCE: [
                CallbackQueryHandler(penerima_source_callback, pattern="^pm_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, penerima_url),
            ],
            STATE_PENERIMA_URL: [MessageHandler(filters.TEXT & ~filters.COMMAND, penerima_url)],
            STATE_PENERIMA_ROW: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, penerima_rows),
                MessageHandler(filters.Document.ALL, penerima_rows),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
        per_message=False,
    )
    app.add_handler(penerima_conv)

    # ── Register bot commands menu ──
    # Shows up when user types / in chat
    async def set_commands(app):
        await app.bot.set_my_commands([
            ("start", "🏠 Menu utama"),
            ("harian", "📋 Input laporan harian"),
            ("mingguan", "📊 Input laporan mingguan"),
            ("mingguan_auto", "🤖 Auto-summary mingguan"),
            ("penerima", "👥 Report penerima manfaat"),
            ("juknis", "📚 Akses JUKNIS lengkap"),
            ("status", "📈 Status operasional hari ini"),
            ("help", "❓ Bantuan"),
            ("cancel", "⛔ Batalkan sesi"),
        ])
        print("✅ Bot commands registered")

    # ── Error handler ──
    app.add_error_handler(error_handler)

    # ── Start polling ──
    print("✅ SPPG Bot is running...")

    # Set bot commands via a one-time job on startup
    async def set_commands_callback(app):
        await app.bot.set_my_commands([
            ("start", "🏠 Menu utama"),
            ("harian", "📋 Input laporan harian"),
            ("mingguan", "📊 Input laporan mingguan"),
            ("mingguan_auto", "🤖 Auto-summary mingguan"),
            ("penerima", "👥 Report penerima manfaat"),
            ("juknis", "📚 Akses JUKNIS lengkap"),
            ("status", "📈 Status operasional hari ini"),
            ("help", "❓ Bantuan"),
            ("cancel", "⛔ Batalkan sesi"),
        ])
        print("✅ Bot commands registered")

    # Use post_init to set commands on startup
    async def post_init(application):
        await set_commands_callback(application)
        return application

    app.post_init = post_init
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()